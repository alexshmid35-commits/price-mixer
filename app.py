#!/usr/bin/env python3
"""
Price Mixer Web — веб-интерфейс для сведения прайсов поставщиков.

Запуск: python3 app.py
Открыть: http://localhost:5001
"""

import json
import logging
import math
import os
import queue
import re
import shutil
import subprocess
import sys
import threading
import time
import urllib.request
from difflib import SequenceMatcher
from functools import wraps
from pathlib import Path

import numpy as np
import pandas as pd
import requests
from flask import (
    Flask,
    Response,
    jsonify,
    redirect,
    render_template,
    request,
    send_file,
    session,
    url_for,
)

from config import cfg
from mixer import (
    build_id_fanout_map,
    consolidate_simple,
    extract_article,
    extract_article_candidates,
    is_trusted_cached_id,
    load_id_cache,
    load_url_cache,
    lookup_catalog_match_details,
    lookup_id_from_catalog_sheet,
    parse_generic_excel,
    resolve_onliner_urls,
    save_id_cache,
)
from price_mixer.api.routes import bp as api_bp
from price_mixer.api.autofill_routes import create_autofill_bp
from price_mixer.api.bulk_id_routes import create_bulk_id_bp
from price_mixer.api.category_management_routes import create_category_management_bp
from price_mixer.api.category_reference_routes import create_category_reference_bp
from price_mixer.api.export_routes import create_export_bp
from price_mixer.api.experimental_noid_routes import create_experimental_noid_bp
from price_mixer.api.id_reporting_routes import create_id_reporting_bp
from price_mixer.api.id_validation_routes import create_id_validation_bp
from price_mixer.api.manual_id_routes import create_manual_id_bp
from price_mixer.api.main_routes import create_main_bp
from price_mixer.api.market_routes import create_market_bp
from price_mixer.api.onliner_db_routes import create_onliner_db_bp
from price_mixer.api.onliner_routes import create_onliner_bp
from price_mixer.api.resolve_routes import create_resolve_bp
from price_mixer.api.review_queue_routes import create_review_queue_bp
from price_mixer.api.settings_routes import create_settings_bp
from price_mixer.api.source_routes import create_source_bp
from price_mixer.logging_config import configure_price_mixer_logging, get_logger
from price_mixer.process_lock import try_acquire_pid_lock
from price_mixer.request_logging import register_request_logging
from price_mixer.runtime_paths import ensure_runtime_directories, get_runtime_paths
from price_mixer.services.api_sources import (
    fetch_api_source_worker as _api_source_fetch_worker,
    get_source_runtime as _api_sources_get_runtime,
)
from price_mixer.services.autofill_workers import (
    make_pc_autofill_status as _autofill_make_pc_status,
    run_tgpc_pc_worker as _autofill_run_tgpc_pc,
    start_autofill_payload as _autofill_start_payload,
    status_payload as _autofill_status_payload,
)
from price_mixer.services.manual_id_actions import reject_iven_match_payload as _reject_iven_match
from price_mixer.services.background_xlsx import create_background_xlsx_worker
from price_mixer.services.category_config import (
    apply_markup_to_df as _category_apply_markup_to_df,
    build_markup_preview_payload as _category_markup_preview_payload,
    calc_rrc_and_no_discount,
    get_category_markup_config,
    load_category_markups,
    load_category_overrides,
    load_manual_category_overrides,
    parse_markup_request as _category_parse_markup_request,
    save_category_markups,
    save_category_overrides,
    save_manual_category_overrides,
    update_markups_for_categories as _category_update_markups,
)
from price_mixer.services.durable_jobs import DurableJobQueue
from price_mixer.services.category_extra_runtime import CategoryExtraRuntime
from price_mixer.services.category_management_runtime import (
    CategoryManagementRuntime,
)
from price_mixer.services.category_state_store import (
    CATEGORY_OVERRIDES_STATE,
    MANUAL_CATEGORY_OVERRIDES_STATE,
    CATEGORY_VISIBILITY_STATE,
    category_state_signature as _category_state_signature,
    load_category_state as _category_state_load,
    save_category_state as _category_state_save,
)
from price_mixer.services.category_reference import (
    build_categories_payload as _category_reference_categories_payload,
    build_category_catalog_payload as _category_reference_catalog_payload,
    build_supplier_categories_payload as _category_reference_supplier_categories_payload,
    build_suppliers_payload as _category_reference_suppliers_payload,
)
from price_mixer.services.category_pipeline import (
    apply_saved_markups_to_df as _category_apply_saved_markups,
    apply_visibility_filter as _category_apply_visibility_filter,
    apply_category_override_to_df as _category_apply_override_to_df,
    build_category_override_items_payload as _category_override_items_payload,
    build_category_preview_items_payload as _category_preview_items_payload,
    ensure_category_column as _category_ensure_column,
    get_effective_category as _category_get_effective,
    GLOBAL_VISIBILITY_KEY,
    load_visibility_map as _category_load_visibility,
    row_category as _category_row_category,
    save_visibility_map as _category_save_visibility,
    update_category_visibility as _category_update_visibility,
)
from price_mixer.services.consolidated_io import (
    delivery_days_from_row as _consolidated_delivery_days_from_row,
    read_consolidated_df as _consolidated_read_df,
    read_consolidated_json_rows as _consolidated_read_json_rows,
    safe_json_value as _consolidated_safe_json_value,
    write_consolidated_df as _consolidated_write_df,
    write_consolidated_json as _consolidated_write_json,
)
from price_mixer.services.consolidated_paging import (
    build_consolidated_page as _build_consolidated_page,
)
from price_mixer.services.export_pipeline import (
    build_preexport_quality_payload as _export_build_preexport_quality_payload,
    dataframe_to_export_dataframe as _export_dataframe_to_xlsx,
    export_google_sheets_payload as _export_google_sheets_payload,
    prepare_consolidated_for_export as _export_prepare_consolidated,
    resolve_service_account_json_path as _export_resolve_service_account_json_path,
)
from price_mixer.services.export_stats import (
    export_category_counts_from_json_rows as _export_stats_category_counts_from_json_rows,
    export_row_count_from_json_rows as _export_stats_row_count_from_json_rows,
    format_category_counts as _export_stats_format_category_counts,
    without_id_category_counts_from_df as _export_stats_without_id_category_counts_from_df,
    without_id_category_counts_from_json_rows as _export_stats_without_id_category_counts_from_json_rows,
)
from price_mixer.services.experimental_noid import ExperimentalNoIdRuntime
from price_mixer.services.export_filters import (
    apply_export_duplicate_id_filter as _export_filter_duplicate_ids,
    apply_export_keep_lowest_price_per_onliner_id as _export_filter_keep_lowest_price,
    apply_export_only_pc_filter as _export_filter_only_pc,
    build_duplicate_onliner_id_issues as _export_build_duplicate_onliner_id_issues,
    is_pc_export_row as _export_is_pc_export_row,
    normalize_supplier_name_list as _export_normalize_supplier_name_list,
)
from price_mixer.services.id_reporting import (
    build_duplicate_onliner_ids_payload as _id_reporting_duplicate_onliner_ids_payload,
    build_id_replace_candidates_payload as _id_reporting_replace_candidates_payload,
)
from price_mixer.services.id_compare_report import (
    build_id_compare_report_df as _id_compare_report_build_df,
)
from price_mixer.services.id_validation import (
    ValidationCancelledError as _IdValidationCancelledError,
    apply_validate_clean_api_result as _id_validation_apply_api_result,
    build_verify_all_start_state as _id_validation_build_verify_start_state,
    build_validate_clean_cancelled_state as _id_validation_clean_cancelled_state,
    build_validate_clean_error_state as _id_validation_clean_error_state,
    build_validate_clean_finish_state as _id_validation_clean_finish_state,
    build_validate_clean_no_column_state as _id_validation_clean_no_column_state,
    build_validate_clean_no_tasks_state as _id_validation_clean_no_tasks_state,
    build_validate_clean_prepare_progress_state as _id_validation_clean_prepare_progress_state,
    collect_id_validation_tasks as _id_validation_collect_tasks,
    populate_api_review_queue_for_cleared_items as _id_validation_populate_api_review_queue,
    populate_db_review_queue_for_cleared_items as _id_validation_populate_db_review_queue,
    run_validate_clean_db_tasks as _id_validation_run_db_tasks,
    sort_verify_result_items as _id_validation_sort_verify_items,
    save_validate_clean_results as _id_validation_save_clean_results,
    verify_onliner_id_row as _id_validation_verify_row,
)
from price_mixer.services.id_validation_runtime import IdValidationRuntime
from price_mixer.services.id_validation_api_worker import run_api_validation_worker
from price_mixer.services.id_validation_db_worker import run_db_validation_worker
from price_mixer.services.id_validation_verify_worker import run_verify_all_worker
from price_mixer.services.isolated_verify_job import IsolatedVerifyJob
from price_mixer.services.validate_clean_analysis import ValidateCleanAnalysisRunner
from price_mixer.services.category_autosort import (
    apply_autosort_items as _autosort_apply_items,
    build_autosort_preview_payload as _autosort_preview_payload,
    predict_openai_category as _autosort_predict_openai_category,
)
from price_mixer.services.market_refresh import (
    get_market_refresh_status_snapshot as _market_refresh_status_snapshot,
    start_market_refresh as _market_refresh_start,
)
from price_mixer.services.manual_id_runtime import ManualIdRuntime
from price_mixer.services.manual_id_store import (
    append_id_change_journal as _manual_store_append_journal,
    is_manually_confirmed_id as _manual_store_is_confirmed_id,
    load_id_change_journal as _manual_store_load_journal,
    load_manual_id_bindings as _manual_store_load_bindings,
    save_id_change_journal as _manual_store_save_journal,
    save_manual_id_bindings as _manual_store_save_bindings,
)
from price_mixer.services.review_queue_store import (
    load_review_queue as _review_store_load_queue,
    save_review_queue as _review_store_save_queue,
)
from price_mixer.services.review_queue import (
    dataframe_id_conflict_for_supplier as _review_queue_dataframe_id_conflict,
    manual_binding_id_conflict as _review_queue_manual_binding_id_conflict,
    match_name_key as _review_queue_service_match_name_key,
    migrate_supplier_scope as _review_queue_migrate_supplier_scope,
    supplier_scoped_key as _review_queue_service_supplier_scoped_key,
    unique_supplier_names as _review_queue_service_supplier_names,
)
from price_mixer.services.review_queue_runtime import ReviewQueueRuntime
from price_mixer.services.main_payloads import (
    build_consolidated_table_rows as _main_build_consolidated_rows,
    build_stats_payload as _main_build_stats_payload,
    empty_stats_payload as _main_empty_stats_payload,
)
from price_mixer.services.ntech_review_queue import (
    no_candidates_review_item as _ntech_no_candidates_review_item,
    queued_review_item as _ntech_queued_review_item,
    skip_review_row as _ntech_skip_review_row,
)
from price_mixer.services.ntech_review_runtime import NTechReviewRuntime
from price_mixer.services.ntech_review_extra import (
    build_generic_category_review_handler as _ntech_build_generic_category_handler,
    build_supplier_laptop_review_handler as _ntech_build_laptop_handler,
    find_review_candidates as _ntech_find_review_candidates,
)
from price_mixer.services.ntech_review_presets import (
    NTECH_CATEGORY_REVIEW_CONFIG as _NTECH_CATEGORY_REVIEW_CONFIG,
    build_core_review_start_kwargs as _ntech_core_review_start_kwargs,
    build_generic_review_start_kwargs as _ntech_generic_review_start_kwargs,
    build_laptop_review_start_kwargs as _ntech_laptop_review_start_kwargs,
)
from price_mixer.services.ntech_review_categories import (
    build_ntech_review_handlers_from_runtime as _build_ntech_review_handlers_from_runtime,
)
from price_mixer.services.catalog_consistency import (
    enforce_catalog_consistency as _catalog_enforce_consistency,
    reconcile_ids_from_catalog as _catalog_reconcile_ids,
)
from price_mixer.services.bulk_id_cleanup import (
    clear_all_nonpc_onliner_ids as _bulk_clear_all_nonpc_onliner_ids,
    clear_invalid_onliner_ids as _bulk_clear_invalid_onliner_ids,
    clear_ntech_duplicate_onliner_ids as _bulk_clear_ntech_duplicate_onliner_ids,
)
from price_mixer.services.onliner_api import onliner_api_get
from price_mixer.services.onliner_b2b import (
    b2b_section_tokens as _b2b_section_tokens,
    get_onliner_b2b_settings,
    invalidate_onliner_b2b_token,
    fetch_market_stats_b2b as _b2b_fetch_market_stats,
    onliner_b2b_fetch_product_positions_export,
    onliner_b2b_get_articles,
    onliner_b2b_get_manufacturers,
    onliner_b2b_get_products,
    onliner_b2b_get_sections,
    onliner_b2b_get_token,
    onliner_b2b_request,
    resolve_catalog_path_for_product as _b2b_resolve_catalog_path_for_product,
    search_candidates as _b2b_search_candidates,
)
from price_mixer.services.onliner_db import (
    catalog_import_worker as _onliner_db_catalog_import_worker,
    db_connection as _onliner_db_connection,
    find_exact_id_for_name as _onliner_db_find_exact_id,
    find_id_for_name as _onliner_db_find_id,
    find_top_candidates as _onliner_db_find_top_candidates,
    get_categories_by_ids as _onliner_db_get_categories_by_ids,
    get_distinct_categories as _onliner_db_get_distinct_categories,
    get_product_by_id as _onliner_db_get_product_by_id,
    import_csv_payload as _onliner_db_import_csv,
    import_gsheet_payload as _onliner_db_import_gsheet,
    import_status_payload as _onliner_db_import_status,
    init_onliner_db as _onliner_db_init,
    populate_from_df as _onliner_db_populate_from_df,
    rebuild_payload as _onliner_db_rebuild,
    search_payload as _onliner_db_search,
    search_tgpc_pc_candidates as _onliner_db_search_tgpc_pc,
    stats_payload as _onliner_db_stats,
    update_categories as _onliner_db_update_categories,
    upsert_product as _onliner_db_upsert_product,
)
from price_mixer.services.onliner_diagnostics import (
    b2b_probe_payload as _onliner_diag_b2b_probe,
    b2b_test_payload as _onliner_diag_b2b_test,
    offers_payload as _onliner_diag_offers,
)
from price_mixer.services.onliner_category_preview import (
    build_onliner_category_preview_payload as _onliner_category_preview_payload,
    collect_onliner_ids as _onliner_category_preview_collect_ids,
)
from price_mixer.services.onliner_market_runtime import OnlinerMarketRuntime
from price_mixer.services.onliner_market import (
    ONLINER_PRODUCT_CACHE_LOCK,
    ONLINER_PRODUCT_CACHE_TTL,
    extract_offer_rows as _extract_offer_rows,
    get_onliner_market_stats_from_cache_only,
    load_onliner_market_cache,
    load_onliner_product_cache,
    save_onliner_product_cache,
)
from price_mixer.services.onliner_search import (
    category_path_hints as _onliner_search_category_path_hints,
    search_candidates as _onliner_search_candidates,
    search_product_by_name as _onliner_search_product_by_name,
    search_product_by_name_deep as _onliner_search_product_by_name_deep,
)
from price_mixer.services.product_normalization import (
    _fallback_category_token,
    build_item_category_key,
    build_item_category_keys,
    count_rows_with_duplicate_onliner_id as _count_rows_with_duplicate_onliner_id,
    count_rows_without_onliner_id as _count_rows_without_onliner_id,
    infer_category,
    normalize_catalog_category_name,
    normalize_consolidated_columns,
    normalize_internal_category_name,
    normalize_name_key as _normalize_name_key,
    normalize_onliner_id,
    round_price_to_90,
)
from price_mixer.services.product_matching import (
    CATEGORY_LOOKUP as _CATEGORY_LOOKUP,
    article_like_tokens as _matching_article_like_tokens,
    calc_name_match as _matching_calc_name_match,
    capacity_tokens as _matching_capacity_tokens,
    color_tokens as _matching_color_tokens,
    extract_gpu_model as _matching_extract_gpu_model,
    extract_product_category as _matching_extract_product_category,
    extract_tgpc_pc_code as _matching_extract_tgpc_pc_code,
    harden_base_verify_result as _matching_harden_base_verify_result,
    important_name_tokens as _matching_important_name_tokens,
    is_color_only_chunk as _matching_is_color_only_chunk,
    is_spec_code as _matching_is_spec_code,
    model_hint_tokens as _matching_model_hint_tokens,
    name_tokens as _matching_name_tokens,
    normalize_compact_name as _matching_normalize_compact_name,
    normalize_match_text as _matching_normalize_match_text,
    ordered_token_hits as _matching_ordered_token_hits,
    paren_chunks as _matching_paren_chunks,
    raw_paren_article_tokens as _matching_raw_paren_article_tokens,
    raw_search_tokens as _matching_raw_search_tokens,
    token_family_match as _matching_token_family_match,
)
from price_mixer.services.processing_pipeline import (
    infer_supplier_from_filename as _processing_infer_supplier_from_filename,
    process_supplier_files as _processing_process_supplier_files,
)
from price_mixer.services.quality_pipeline import (
    apply_quality_visibility_filter as _quality_apply_visibility_filter,
)
from price_mixer.services.sorting_reparse_client import (
    parser_error_message as _sorting_reparse_error_message,
    response_json_payload as _sorting_reparse_json_payload,
)
from price_mixer.services.source_runtime import SourceRuntime
from price_mixer.services.review_candidates import (
    board_brand_model_key as _review_board_brand_model_key,
    case_brand_model_key as _review_case_brand_model_key,
    cooler_brand_model_key as _review_cooler_brand_model_key,
    cpu_brand_model_key as _review_cpu_brand_model_key,
    find_board_review_candidates as _review_find_board_candidates,
    find_case_review_candidates as _review_find_case_candidates,
    find_cooler_review_candidates as _review_find_cooler_candidates,
    find_cpu_review_candidates as _review_find_cpu_candidates,
    find_gpu_review_candidates as _review_find_gpu_candidates,
    find_hdd_review_candidates as _review_find_hdd_candidates,
    find_monitor_review_candidates as _review_find_monitor_candidates,
    find_peripheral_review_candidates as _review_find_peripheral_candidates,
    find_printer_review_candidates as _review_find_printer_candidates,
    find_psu_review_candidates as _review_find_psu_candidates,
    find_ram_review_candidates as _review_find_ram_candidates,
    find_ssd_review_candidates as _review_find_ssd_candidates,
    gpu_brand_model_key as _review_gpu_brand_model_key,
    hdd_brand_model_key as _review_hdd_brand_model_key,
    looks_like_case_name as _review_looks_like_case_name,
    looks_like_cooler_name as _review_looks_like_cooler_name,
    looks_like_hdd_name as _review_looks_like_hdd_name,
    looks_like_liquid_cpu_cooling_name as _review_looks_like_liquid_cpu_cooling_name,
    looks_like_peripheral_name as _review_looks_like_peripheral_name,
    looks_like_printer_or_mfp_name as _review_looks_like_printer_or_mfp_name,
    looks_like_cpu_name as _review_looks_like_cpu_name,
    monitor_brand_model_key as _review_monitor_brand_model_key,
    printer_mfp_brand_model_key as _review_printer_mfp_brand_model_key,
    psu_brand_model_key as _review_psu_brand_model_key,
    ram_brand_model_key as _review_ram_brand_model_key,
    ssd_brand_model_key as _review_ssd_brand_model_key,
)
from price_mixer.services.resolve_pipeline import (
    make_resolve_status as _resolve_make_status,
    resolve_status_snapshot as _resolve_status_snapshot,
    start_resolve_payload as _resolve_start_payload,
)
from price_mixer.services.supplier_snapshots import (
    append_api_fetch_history,
    get_api_fetch_history,
    load_session_supplier_diff,
    save_session_supplier_diff,
)
from price_mixer.services.upload_sessions import (
    cleanup_old_uploads as _upload_sessions_cleanup_old,
    create_session_dir as _upload_sessions_create_dir,
    maybe_cleanup_old_uploads as _upload_sessions_maybe_cleanup_old,
)
from price_mixer.services.upload_files import build_upload_file_entries as _upload_files_build_entries
from price_mixer.settings import (
    _coerce_bool,
    _coerce_float,
    _coerce_int,
    get_onliner_api_max_workers,
    load_app_settings,
    load_auto_refresh_settings,
    load_onliner_api_settings,
    save_app_settings as _save_app_settings_base,
    save_auto_refresh_settings,
    save_onliner_api_settings,
)
from price_mixer.state_store import load_dict, save_dict, save_json_atomic
from price_mixer.web_helpers import active_session_dir, basic_auth_matches, resolve_session_dir

# Глобальный прогресс резолвинга
resolve_status = _resolve_make_status()
verify_all_ids_status = {
    "running": False,
    "total": 0,
    "done": 0,
    "matched": 0,
    "mismatched": 0,
    "errors": 0,
    "items": [],
    "report_items": [],
    "started_at": 0,
    "finished_at": 0,
    "message": "",
}
VERIFY_ALL_IDS_LOCK = threading.RLock()
VERIFY_ALL_IDS_STATUS_WRITER = None
VERIFY_ALL_IDS_JOB = IsolatedVerifyJob()
validate_clean_ids_status = {
    "running": False,
    "total": 0,
    "done": 0,
    "confirmed": 0,
    "cleared": 0,
    "queued": 0,
    "errors": 0,
    "mode": "api",
    "mode_label": "Onliner API",
    "skipped_label": "Пропуск = API не ответил, ID не меняли.",
    "started_at": 0,
    "finished_at": 0,
    "message": "",
}
VALIDATE_CLEAN_IDS_LOCK = threading.RLock()
VALIDATE_CLEAN_ANALYSIS_RUNNER = ValidateCleanAnalysisRunner()
VALIDATE_CLEAN_CANCEL_EVENT = threading.Event()
autofill_ntech_pc_status = {
    "running": False,
    "total": 0,
    "done": 0,
    "applied": 0,
    "skipped": 0,
    "percent": 0,
    "items": [],
    "message": "",
}
AUTOFILL_NTECH_PC_LOCK = threading.RLock()
autofill_iven_pc_status = {
    "running": False,
    "total": 0,
    "done": 0,
    "applied": 0,
    "skipped": 0,
    "percent": 0,
    "items": [],
    "message": "",
}
AUTOFILL_IVEN_PC_LOCK = threading.RLock()
id_review_status = {
    "running": False,
    "total": 0,
    "done": 0,
    "applied": 0,
    "skipped": 0,
    "percent": 0,
    "started_at": 0,
    "finished_at": 0,
    "message": "",
    "matches": [],   # [{name, matched_name, score, id, url, source}]
    "no_match": [],  # [{name}] — не нашли
    "report_mode": "iven",
    "report_title": "Отчёт проверки ID",
    "report_subtitle": "Кандидаты для ручного подтверждения",
}
ID_REVIEW_STATUS_LOCK = threading.RLock()
AUTO_REFRESH_MAX_IDS = 1200

app = Flask(__name__)
app.secret_key = cfg.get_flask_secret_key()
app.config["MAX_CONTENT_LENGTH"] = 200 * 1024 * 1024
configure_price_mixer_logging()
APP_LOGGER = get_logger("price_mixer.app")
register_request_logging(app, APP_LOGGER)
app.register_blueprint(api_bp)
RUNTIME_PATHS = ensure_runtime_directories(get_runtime_paths())
PRICE_DATA_MUTATION_LOCK = threading.RLock()
BACKGROUND_XLSX_WORKER = create_background_xlsx_worker()
DURABLE_JOB_QUEUE = getattr(BACKGROUND_XLSX_WORKER, "queue", None)
BACKGROUND_STARTED = False
BACKGROUND_LOCK_HANDLE = None
EXPERIMENTAL_NOID_RUNTIME = None
EXPERIMENTAL_NOID_RUNTIME_LOCK = threading.Lock()
LAST_ACTIVE_SESSION_DIR = None
QUALITY_STATS_CACHE = {}
GOOGLE_EXPORT_DF_CACHE = {}
LAST_UPLOAD_CLEANUP_TS = 0


def _serialized_price_mutation(func):
    @wraps(func)
    def _wrapped(*args, **kwargs):
        with PRICE_DATA_MUTATION_LOCK:
            return func(*args, **kwargs)

    return _wrapped

UPLOAD_DIR = RUNTIME_PATHS.uploads_dir
UPLOAD_DIR.mkdir(exist_ok=True)
UPLOAD_KEEP_LAST_SESSIONS = 20
UPLOAD_KEEP_DAYS = 7
UPLOAD_KEEP_API_FETCH_HOURS = 12

OPENAI_API_KEY = os.getenv("OPENAI_API_KEY", "").strip()
OPENAI_AUTOSORT_MODEL = os.getenv("OPENAI_AUTOSORT_MODEL", "gpt-4o-mini").strip() or "gpt-4o-mini"
OPENAI_AUTOSORT_TIMEOUT_SEC = 9
OPENAI_AUTOSORT_MAX_ITEMS = 320
OPENAI_AUTOSORT_MAX_WORKERS = 10
AI_CATEGORY_CACHE = {}
AI_CATEGORY_CACHE_LOCK = threading.Lock()


def _resolve_session_dir(session_id):
    return resolve_session_dir(UPLOAD_DIR, session_id)


def get_active_session_dir():
    """Return the active upload session directory, constrained to UPLOAD_DIR."""
    candidate = active_session_dir(UPLOAD_DIR, session)
    return str(candidate) if candidate else None


def _ensure_background_workers():
    global BACKGROUND_LOCK_HANDLE, BACKGROUND_STARTED
    if BACKGROUND_STARTED:
        return
    BACKGROUND_LOCK_HANDLE = try_acquire_pid_lock(UPLOAD_DIR / "_background_workers.pid")
    BACKGROUND_STARTED = True
    if BACKGROUND_LOCK_HANDLE is None:
        return
    threading.Thread(target=_auto_market_refresh_loop, daemon=True).start()


def _cleanup_old_uploads(exclude_dirs=None):
    return _upload_sessions_cleanup_old(
        UPLOAD_DIR,
        load_settings=load_app_settings,
        exclude_dirs=exclude_dirs,
        keep_last_sessions_default=UPLOAD_KEEP_LAST_SESSIONS,
        keep_days_default=UPLOAD_KEEP_DAYS,
        keep_api_fetch_hours_default=UPLOAD_KEEP_API_FETCH_HOURS,
    )


def _maybe_cleanup_old_uploads(exclude_dirs=None, min_interval_sec=1800):
    global LAST_UPLOAD_CLEANUP_TS
    result, LAST_UPLOAD_CLEANUP_TS = _upload_sessions_maybe_cleanup_old(
        cleanup=lambda: _cleanup_old_uploads(exclude_dirs=exclude_dirs),
        last_cleanup_ts=LAST_UPLOAD_CLEANUP_TS,
        min_interval_sec=min_interval_sec,
    )
    return result


@app.before_request
def _startup_background_workers():
    global LAST_ACTIVE_SESSION_DIR
    _ensure_background_workers()
    sdir = get_active_session_dir()
    if sdir:
        LAST_ACTIVE_SESSION_DIR = str(sdir)
    exclude = [LAST_ACTIVE_SESSION_DIR, sdir]
    _maybe_cleanup_old_uploads(exclude_dirs=exclude)

# ============================================================
# HTML ШАБЛОНЫ

# ============================================================
# ROUTES
# ============================================================

def index():
    error = request.args.get("error")
    response = app.make_response(render_template("upload.html", error=error))
    response.headers["Cache-Control"] = "no-store, no-cache, must-revalidate, max-age=0"
    response.headers["Pragma"] = "no-cache"
    response.headers["Expires"] = "0"
    return response


def result_page():
    sid = str(request.args.get("sid", "") or "").strip()
    app_settings = load_app_settings()
    session_dir = get_active_session_dir()
    if sid:
        candidate_dir = _resolve_session_dir(sid)
        if candidate_dir is None:
            return redirect(url_for("main_api.index", error="Недопустимый идентификатор сессии"))
        candidate_file = candidate_dir / "consolidated_price.xlsx"
        if _has_consolidated_session_file(candidate_dir):
            session_dir = str(candidate_dir)
            _finalize_processed_session(sid, candidate_dir, candidate_file)
    if not session_dir:
        return redirect(url_for("main_api.index", error="Нет активного прайса"))
    if not _has_consolidated_session_file(session_dir):
        return redirect(url_for("main_api.index", error="Файл результата не найден"))
    df = read_consolidated_json_fast_df(session_dir)
    total_suppliers = len(set(str(v).strip() for v in df.get("Поставщик", pd.Series(dtype=str)).tolist() if str(v).strip()))
    without_id = _count_rows_without_onliner_id(df)
    duplicate_id_rows = _count_rows_with_duplicate_onliner_id(df)
    export_rows = _export_row_count_for_session(session_dir)
    export_category_counts = _export_category_counts_for_session(session_dir)
    without_id_category_counts = _without_id_category_counts_from_df(df)
    hidden_category_counts = _hidden_category_counts_for_session(session_dir)
    hidden_rows = int(sum(int(item.get("count", 0) or 0) for item in hidden_category_counts))
    with_id = len(df) - without_id
    stats = {
        "total": len(df),
        "suppliers": total_suppliers,
        "consolidated": len(df),
        "matched": with_id,
        "with_id": with_id,
        "without_id": without_id,
        "duplicate_id_rows": duplicate_id_rows,
        "export_rows": export_rows,
        "export_category_counts": export_category_counts,
        "without_id_category_counts": without_id_category_counts,
        "hidden_rows": hidden_rows,
        "hidden_category_counts": hidden_category_counts,
        "show_checks_block": _coerce_bool((((app_settings or {}).get("ui") or {}).get("show_checks_block", True)), default=True),
        "snapshot_diff": load_session_supplier_diff(session_dir),
    }
    return render_template("result.html", stats=stats)




@app.before_request
def require_basic_auth():
    """Require HTTP Basic Auth for all routes except health/version."""
    exempt_paths = {
        "/api/health",
        "/api/version",
        # Read-only endpoint for CRM product-name binding. It exposes only
        # local Onliner catalog search results and does not mutate server data.
        "/api/onliner-db-search",
        "/api/onliner-db-stats",
    }
    if request.path in exempt_paths:
        return None
    try:
        admin_username = str(app.config.get("ADMIN_USERNAME", cfg.admin_username))
        admin_password = str(app.config.get("ADMIN_PASSWORD") or cfg.require_admin_password())
    except RuntimeError as exc:
        return Response(str(exc) + "\n", 500)
    auth = request.authorization
    if not basic_auth_matches(auth, admin_username, admin_password):
        return Response(
            "Authentication required\n",
            401,
            {"WWW-Authenticate": 'Basic realm="Price Mixer"'},
        )
@app.after_request
def add_no_cache_headers(response):
    try:
        response.headers["Cache-Control"] = "no-store, no-cache, must-revalidate, max-age=0"
        response.headers["Pragma"] = "no-cache"
        response.headers["Expires"] = "0"
    except Exception:
        pass
    return response


def get_article_from_name(name):
    """Извлечь артикул из названия."""
    if not name:
        return ""
    article = extract_article(name)
    if article:
        return article
    # No fallback to name fragments: it causes cache-key collisions
    # for similar long names (notably motherboards) and can propagate one ID.
    return ""


def _is_generic_cpu_cache_key(key):
    raw = str(key or "").strip()
    if not raw:
        return True
    low = raw.lower()
    compact = re.sub(r"[\s_]+", "", low)
    generic_patterns = [
        r"^socket[-\s]?[a-z0-9-]+$",
        r"^lga[-\s]?\d{3,5}$",
        r"^am\d+$",
        r"^fm\d+$",
        r"^tr\d+[a-z]*$",
        r"^s\d{3,5}$",
    ]
    for pattern in generic_patterns:
        if re.match(pattern, low, flags=re.IGNORECASE) or re.match(pattern, compact, flags=re.IGNORECASE):
            return True
    return False


def _is_generic_id_cache_key(key):
    raw = str(key or "").strip()
    if not raw:
        return True
    low = raw.lower()
    compact = re.sub(r"[^a-z0-9]+", "", low)
    generic_patterns = [
        r"^soc[-\s]?\d{3,5}[a-z]*$",
        r"^socket[-\s]?[a-z0-9-]{3,}$",
        r"^lga[-\s]?\d{3,5}[a-z]*$",
        r"^am\d+[a-z]*$",
        r"^fm\d+[a-z]*$",
        r"^tr\d+[a-z]*$",
    ]
    for pattern in generic_patterns:
        if re.match(pattern, low, flags=re.IGNORECASE) or re.match(pattern, compact, flags=re.IGNORECASE):
            return True
    return False


def _get_id_cache_key_for_name(name):
    article = str(get_article_from_name(name) or "").strip()
    if not article:
        return ""
    name_low = str(name or "").strip().lower()
    is_cpu_like = any(token in name_low for token in ["процессор", "intel", "amd", "ryzen", "xeon", "celeron", "pentium", "athlon", "core i"])
    if _is_generic_id_cache_key(article):
        return ""
    if is_cpu_like and _is_generic_cpu_cache_key(article):
        return ""
    return article


def _iven_pc_cache_key_for_name(name):
    code = _extract_iven_pc_code(name)
    return f"iven_pc:{code}" if code else ""


def _iven_pc_series_cache_key_for_name(name):
    code = _extract_iven_pc_code(name)
    series = _extract_iven_pc_series(name)
    return f"iven_pc:{series}:{code}" if code and series else ""


def _id_cache_keys_for_iven_pc_name(name):
    return [
        _iven_pc_cache_key_for_name(name),
        _iven_pc_series_cache_key_for_name(name),
    ]


def _supplier_key_token(supplier_name):
    token = str(supplier_name or "").strip().lower()
    for old, new in ((" ", "_"), ("-", "_")):
        token = token.replace(old, new)
    while "__" in token:
        token = token.replace("__", "_")
    return token.strip("_")


def _manual_binding_scoped_key(name_key, supplier_name):
    base_key = str(name_key or "").strip()
    supplier_token = _supplier_key_token(supplier_name)
    return f"supplier:{supplier_token}:{base_key}" if base_key and supplier_token else ""


def _manual_binding_keys_for_name(name, supplier_name=None):
    if _is_iven_pc_name(name):
        code_key = _iven_pc_cache_key_for_name(name)
        series_key = _iven_pc_series_cache_key_for_name(name)
        base_keys = [code_key, series_key, _normalize_name_key(name)]
    else:
        base_keys = [_normalize_name_key(name)]
    keys = []
    if supplier_name:
        keys.extend(_manual_binding_scoped_key(key, supplier_name) for key in base_keys)
    keys.extend(base_keys)
    out = []
    seen = set()
    for key in keys:
        key = str(key or "").strip()
        if key and key not in seen:
            seen.add(key)
            out.append(key)
    return out


def _iven_pc_catalog_code_for_id(onliner_id):
    oid = normalize_onliner_id(onliner_id)
    if not oid:
        return ""
    try:
        product = db_get_product_by_id(oid)
    except Exception:
        product = None
    if not isinstance(product, dict):
        return ""
    return _extract_iven_pc_code(product.get("name", ""))


def _iven_pc_onliner_id_matches_name(name, onliner_id):
    local_code = _extract_iven_pc_code(name)
    catalog_code = _iven_pc_catalog_code_for_id(onliner_id)
    return bool(local_code and catalog_code and local_code == catalog_code)


def _iven_pc_onliner_id_mismatch_known(name, onliner_id):
    local_code = _extract_iven_pc_code(name)
    catalog_code = _iven_pc_catalog_code_for_id(onliner_id)
    return bool(local_code and catalog_code and local_code != catalog_code)


def _manual_binding_matches_name(name, manual):
    if not _is_iven_pc_name(name):
        return True
    if bool((manual or {}).get("blocked", False)):
        return True
    return _iven_pc_onliner_id_matches_name(name, (manual or {}).get("id", ""))


_IVEN_LAPTOP_EXCLUDE_RE = re.compile(
    r"\b(?:сумк\w*|чехл\w*|рюкзак\w*|подставк\w*|столик\w*|кулер\w*|"
    r"охлаждающ\w*|заряд\w*|зарядн\w*|адаптер\w*|блок\s+питания|"
    r"аккумулятор\w*|кабел\w*|матриц\w*|клавиатур\w*|петл\w*|"
    r"рамк\w*|шлейф\w*|док[-\s]?станц\w*|докинг\w*|мыш\w*)\b",
    flags=re.IGNORECASE,
)


def _looks_like_laptop_text(text):
    raw = str(text or "").strip().lower()
    if not raw or _IVEN_LAPTOP_EXCLUDE_RE.search(raw):
        return False
    return bool(re.search(r"\b(?:ноутбук|laptop|notebook|ultrabook)\b", raw, flags=re.IGNORECASE))


def _is_iven_laptop_name(text, category=""):
    raw = str(text or "").strip().lower()
    cat = str(category or "").strip().lower()
    if not raw:
        return False
    if _IVEN_LAPTOP_EXCLUDE_RE.search(raw):
        return False
    return _looks_like_laptop_text(raw) or "ноутбук" in cat or "laptop" in cat or "notebook" in cat


def _is_iven_laptop_candidate(name, url=""):
    raw = str(name or "").strip().lower()
    link = str(url or "").strip().lower()
    if _looks_like_laptop_text(raw):
        return True
    return bool(link and any(part in link for part in ("/notebook/", "/laptop/", "/ultrabook/")))


def _is_tradex_laptop_name(text, category=""):
    return _is_iven_laptop_name(text, category)


def _is_tradex_laptop_candidate(name, url=""):
    return _is_iven_laptop_candidate(name, url)


def _allow_manual_binding_for_supplier(supplier_name, name, category=""):
    # Supplier-scoped records are safe for every category. The lookup helper
    # already rejects legacy/global records that do not name this supplier.
    return True


def _row_price_for_dedupe(row):
    try:
        value = pd.to_numeric(row.get("Цена", np.nan), errors="coerce")
        if pd.isna(value):
            return float("inf")
        return float(value)
    except Exception:
        return float("inf")


def _clear_duplicate_onliner_ids_for_suppliers(df, supplier_names):
    if df is None or df.empty or "OnlinerID" not in df.columns:
        return 0
    supplier_lookup = {
        str(name or "").strip().upper()
        for name in (supplier_names or [])
        if str(name or "").strip()
    }
    if not supplier_lookup:
        return 0
    groups = {}
    for idx, row in df.iterrows():
        supplier = str(row.get("Поставщик", "") or "").strip().upper()
        if supplier not in supplier_lookup:
            continue
        oid = normalize_onliner_id(row.get("OnlinerID", ""))
        if not oid:
            continue
        groups.setdefault((supplier, oid), []).append(idx)

    cleared = 0
    for _key, indices in groups.items():
        if len(indices) <= 1:
            continue
        keep_idx = min(indices, key=lambda row_idx: (_row_price_for_dedupe(df.loc[row_idx]), int(row_idx)))
        for row_idx in indices:
            if row_idx == keep_idx:
                continue
            df.at[row_idx, "OnlinerID"] = ""
            if "Ссылка" in df.columns:
                df.at[row_idx, "Ссылка"] = ""
            cleared += 1
    return cleared


def _lookup_manual_binding_for_name(manual_bindings, name, supplier_name=None):
    if not isinstance(manual_bindings, dict):
        return None
    for key in _manual_binding_keys_for_name(name, supplier_name):
        manual = manual_bindings.get(key)
        if not isinstance(manual, dict) or not _manual_binding_matches_name(name, manual):
            continue
        if supplier_name and not str(key).startswith("supplier:"):
            manual_suppliers = _manual_binding_supplier_names(manual)
            if not manual_suppliers or str(supplier_name).strip().upper() not in _supplier_name_lookup(manual_suppliers):
                continue
        return manual
    return None


def _supplier_name_lookup(supplier_names):
    return {
        str(supplier or "").strip().upper()
        for supplier in (supplier_names or [])
        if str(supplier or "").strip()
    }


def _row_matches_supplier_names(row, supplier_names):
    suppliers = _supplier_name_lookup(supplier_names)
    if not suppliers:
        return True
    supplier = str(row.get("Поставщик", "") or "").strip().upper()
    return supplier in suppliers


def _manual_binding_supplier_names(manual):
    if not isinstance(manual, dict):
        return []
    raw_suppliers = manual.get("suppliers", None)
    if raw_suppliers is None:
        raw_suppliers = manual.get("supplier", "")
    if isinstance(raw_suppliers, str):
        items = [raw_suppliers]
    elif isinstance(raw_suppliers, (list, tuple, set)):
        items = list(raw_suppliers)
    else:
        items = []
    return [str(item or "").strip() for item in items if str(item or "").strip()]


def _apply_manual_bindings_to_consolidated_df(df):
    if df is None or df.empty or "Название" not in df.columns or "OnlinerID" not in df.columns:
        return df
    try:
        manual_bindings = load_manual_id_bindings()
        manual_bindings, aliases_changed = _expand_iven_pc_manual_aliases(manual_bindings)
        if aliases_changed:
            save_manual_id_bindings(manual_bindings)
    except Exception as exc:
        APP_LOGGER.warning("manual bindings overlay skipped: %s", exc)
        return df
    if not isinstance(manual_bindings, dict) or not manual_bindings:
        return df

    out = df.copy()
    if "Ссылка" not in out.columns:
        out["Ссылка"] = ""
    out["OnlinerID"] = out["OnlinerID"].astype("object")
    out["Ссылка"] = out["Ссылка"].astype("object")
    changed = 0
    for idx, row in out.iterrows():
        name = row.get("Название", "")
        supplier_name = str(row.get("Поставщик", "") or "").strip().upper()
        category = row.get("Категория", "")
        if not _allow_manual_binding_for_supplier(supplier_name, name, category):
            continue
        manual = _lookup_manual_binding_for_name(manual_bindings, name, supplier_name)
        if not isinstance(manual, dict):
            continue
        manual_suppliers = _manual_binding_supplier_names(manual)
        if manual_suppliers and supplier_name not in _supplier_name_lookup(manual_suppliers):
            continue
        if bool(manual.get("blocked", False)):
            if normalize_onliner_id(row.get("OnlinerID", "")):
                out.at[idx, "OnlinerID"] = ""
                out.at[idx, "Ссылка"] = ""
                changed += 1
            continue
        oid = normalize_onliner_id(manual.get("id", ""))
        if not oid:
            continue
        old_oid = normalize_onliner_id(row.get("OnlinerID", ""))
        if old_oid != oid:
            out.at[idx, "OnlinerID"] = oid
            changed += 1
        url = str(manual.get("url", "") or "").strip()
        if url and str(row.get("Ссылка", "") or "").strip() != url:
            out.at[idx, "Ссылка"] = url
            changed += 1
    if changed:
        _clear_duplicate_onliner_ids_for_suppliers(out, ["IVEN"])
    return out


def _expand_iven_pc_manual_aliases(manual_bindings):
    if not isinstance(manual_bindings, dict):
        return {}, False
    expanded = dict(manual_bindings)
    changed = False
    for raw_key, record in list(manual_bindings.items()):
        if not isinstance(record, dict):
            continue
        key_text = str(raw_key or "").strip()
        supplier_token = ""
        if key_text.startswith("supplier:"):
            scoped_parts = key_text.split(":", 2)
            if len(scoped_parts) != 3 or not scoped_parts[1] or not scoped_parts[2]:
                continue
            supplier_token = scoped_parts[1]
            key_text = scoped_parts[2]
        if not _is_iven_pc_name(key_text):
            continue
        aliases = []
        for alias in _manual_binding_keys_for_name(key_text):
            if supplier_token:
                alias = _manual_binding_scoped_key(alias, supplier_token)
            if alias:
                aliases.append(alias)
        if aliases and all(alias in expanded for alias in aliases):
            continue
        if not _manual_binding_matches_name(key_text, record):
            continue
        for alias in aliases:
            if alias and alias not in expanded:
                expanded[alias] = dict(record)
                changed = True
    return expanded, changed


def _sanitize_id_cache(cache):
    if not isinstance(cache, dict):
        return {}, False
    cleaned = {}
    changed = False
    for key, value in cache.items():
        if _is_generic_cpu_cache_key(key) or _is_generic_id_cache_key(key):
            changed = True
            continue
        cleaned[key] = value
    return cleaned, changed


CATEGORY_PRIORITY = [
    "Процессор",
    "Кулер",
    "Кулеры",
    "Охлаждение",
    "Воздуходувки",
    "Материнская плата",
    "Оперативная память",
    "SSD",
    "Жесткий диск",
    "Видеокарта",
    "Блок питания",
    "Корпус",
    "Монитор",
    "Кронштейны",
    "Компьютеры",
    "Системные блоки",
    "Моноблоки",
    "Ноутбук",
    "Принтеры",
    "Картриджи",
    "Клавиатура",
    "Мышь",
    "Наушники",
    "Акустика",
    "Портативные колонки",
    "Саундбары",
    "Накопители USB",
    "Внешние накопители",
    "Карты памяти",
    "Кабели и переходники",
    "Сеть",
    "Wi-Fi роутеры",
    "Коммутаторы",
    "Точки доступа Wi-Fi",
    "Беспроводные адаптеры",
    "Сетевые адаптеры",
    "DSL-модемы",
    "USB-хабы",
    "ИБП",
    "Стабилизаторы и сетевые фильтры",
    "Комплекты периферии",
    "Веб-камеры",
    "Коврики для мыши",
    "Боксы для накопителей",
    "Картридеры",
    "Оптические приводы",
    "Звуковые карты",
    "Аксессуары для наушников",
    "Термопасты и термопрокладки",
    "Сумки и чехлы для ноутбуков",
    "Смартфоны",
    "Телевизоры",
    "Планшеты",
    "Умные часы",
    "Умный дом",
    "Игровые приставки",
    "Моддинг ПК",
    "Чистящие средства",
]

EXPORT_ONLINER_STRUCTURE_CATEGORIES = [
    *CATEGORY_PRIORITY,
    "Системный блок",
    "Мини-ПК и одноплатные компьютеры",
    "Принтер и МФУ",
    "МФУ",
    "Бумага и материалы для печати",
    "Фотобумага",
    "Сканеры",
    "Сканеры штрих-кодов",
    "Ламинаторы",
    "Расходные материалы для ламинаторов и брошюровщиков",
    "Микрофоны",
    "Спикерфоны",
    "AV-ресиверы и усилители",
    "Периферия",
    "Оптические диски",
    "Сетевые накопители (NAS)",
    "Сети по электропроводке (Powerline)",
    "Проводные телефоны",
    "Радиотелефоны DECT",
    "Антенны беспроводной связи",
    "FM-модуляторы",
    "Аккумуляторы для ИБП",
    "Батарейки, аккумуляторы, зарядные",
    "Внешние аккумуляторы",
    "Зарядные устройства",
    "Пуско-зарядные устройства",
    "Розетки, выключатели",
    "Электрические щиты",
    "Кабельный крепеж",
    "IP-камеры",
    "Камеры CCTV",
    "Видеодомофоны",
    "Видеорегистраторы",
    "Автомобильные видеорегистраторы",
    "Карты видеозахвата",
    "Информационные панели",
    "Проекторы",
    "Проекционные экраны",
    "Умные часы и браслеты",
    "Игровые контроллеры и аксессуары",
    "Подставки для ноутбуков, телефонов, планшетов",
    "Рюкзаки",
    "Автомобильные держатели",
    "Аксессуары для салона автомобиля",
    "Аксессуары",
    "Аксессуары для оргтехники",
    "Графические планшеты",
    "Шредеры",
    "Офисные кресла",
    "Детские парты, столы, стулья",
    "Обогреватели",
    "Медиаплееры и ТВ-приставки",
    "Плееры",
    "Приемники цифрового ТВ",
    "Пульты ДУ",
    "Радионяни и видеоняни",
    "Радиоприемники",
    "Мультиметры",
    "Наборы инструментов",
    "Строительный, слесарный, монтажный инструмент",
    "Расходные материалы и аксессуары для 3D-печати",
]

EXPORT_MANDATORY_CATEGORY_PREFIX_EXCLUDES = [
    "Требует сортировки",
]

EXPORT_MANDATORY_NAME_EXCLUDES = [
    "патрон",
    "milwaukee",
    "p.i.t",
]

DEFAULT_CATEGORY_VISIBILITY_FILE = RUNTIME_PATHS.state_file("category_visibility.json")
CATEGORY_VISIBILITY_FILE = DEFAULT_CATEGORY_VISIBILITY_FILE
CATEGORY_VISIBILITY_LOCK = threading.RLock()
CATEGORY_OVERRIDE_LOCK = threading.RLock()
CORRECTED_JSON_ROWS_CACHE = {}
CORRECTED_JSON_ROWS_CACHE_LOCK = threading.RLock()
CORRECTED_JSON_ROWS_CACHE_LIMIT = 8
ID_REPLACE_QUERY_CACHE_TTL = 3600
ID_REPLACE_QUERY_CACHE = {}
ID_REPLACE_QUERY_CACHE_LOCK = threading.RLock()
ID_REPLACE_QUERY_CACHE_VERSION = "v7"
CATEGORY_VISIBILITY_SUCCESSORS = {
    "Кулер": {"Кулеры"},
    "Принтер и МФУ": {"Принтеры"},
    "Системный блок": {"Компьютеры", "Системные блоки", "Моноблоки"},
    "КАРТРИДЖ": {"Картриджи"},
    "СУМКА": {"Сумки и чехлы для ноутбуков"},
}

UI_CATEGORY_ALIASES = {
    "аккумуляторная": "АККУМУЛЯТОР",
    "кулер": "Кулеры",
    "принтер и мфу": "Принтеры",
    "системный блок": "Компьютеры",
    "системные блоки": "Компьютеры",
    "картридж": "Картриджи",
    "сумка": "Сумки и чехлы для ноутбуков",
    "термопасты и термопрокладки": "Кулеры",
}

SORTING_REVIEW_PREFIX = "Требует сортировки · родитель: "


def _canonical_ui_category_name(category):
    """Keep result-page categories aligned with the visible Onliner structure."""
    text = normalize_internal_category_name(category)
    if not text:
        return ""
    return UI_CATEGORY_ALIASES.get(text.casefold(), text)


def _canonical_visible_onliner_category_name(category):
    """Preserve already-visible Onliner titles, normalize only raw catalog codes."""
    text = normalize_internal_category_name(category)
    if not text:
        return ""
    if re.search(r"[А-Яа-яЁё]", text) and not _looks_like_raw_supplier_category(text):
        return _canonical_ui_category_name(text)
    return _canonical_ui_category_name(normalize_catalog_category_name(text))


def _sorting_review_category(parent_category):
    parent = str(parent_category or "").strip()
    if parent.startswith(SORTING_REVIEW_PREFIX):
        parent = parent[len(SORTING_REVIEW_PREFIX):].strip()
    parent = _canonical_ui_category_name(parent) or "Без категории"
    return SORTING_REVIEW_PREFIX + parent


def _is_sorting_review_category(category):
    return str(category or "").strip().startswith(SORTING_REVIEW_PREFIX)


def _native_catalog_category_for_product(raw_category, product_name):
    raw = str(raw_category or "").strip()
    name = str(product_name or "").strip()
    if raw.casefold() == "thermal" and re.search(r"\bтермо[\s-]?принтер\b|\bпринтер\b", name, flags=re.IGNORECASE):
        return "Принтеры"
    category = normalize_catalog_category_name(raw)
    if category != "Кабели и переходники":
        return category
    inferred = normalize_catalog_category_name(infer_category(name))
    if _should_repair_catalog_category(category, inferred):
        return inferred
    return category


def _should_repair_catalog_category(catalog_category, inferred_category):
    catalog = str(catalog_category or "").strip()
    inferred = str(inferred_category or "").strip()
    if not catalog or not inferred or catalog == inferred:
        return False
    if catalog == "Кабели и переходники" and inferred in {
        "Охлаждение",
        "Кронштейны",
        "Монитор",
        "Периферия",
        "Клавиатура",
        "Мышь",
        "Наушники",
        "Сеть",
        "Сверла и буры",
        "Строительный, слесарный, монтажный инструмент",
        "Наборы инструментов",
        "Наборы электроинструмента",
    }:
        return True
    return False


def _category_sort_key(category_name):
    if category_name in CATEGORY_PRIORITY:
        return (0, CATEGORY_PRIORITY.index(category_name), category_name)
    return (1, 999, category_name.lower())


def _canonical_supplier_name(supplier):
    text = str(supplier or "").strip()
    compact = text.lower().replace("-", "").replace("_", "").replace(" ", "")
    if compact == "ntech":
        return "N-Tech"
    if compact == "iven":
        return "IVEN"
    if compact == "ivenzakaz":
        return "IVEN_zakaz"
    if compact == "tradex":
        return "Tradex"
    if compact == "tgpc":
        return "TGPC"
    return text


def _category_override_for_row(row, overrides):
    for key in build_item_category_keys(row):
        category = _canonical_ui_category_name((overrides or {}).get(key, ""))
        if category:
            return category
    return ""


def _repair_saved_category_for_product(category, product_name):
    current = _canonical_ui_category_name(category)
    if not current:
        return ""
    if current != "Кабели и переходники":
        return current
    inferred = _canonical_ui_category_name(normalize_catalog_category_name(infer_category(product_name)))
    if _should_repair_catalog_category(current, inferred):
        return inferred
    return current


def _strong_inferred_category_for_product(product_name):
    inferred = _canonical_ui_category_name(normalize_catalog_category_name(infer_category(product_name)))
    fallback = str(_fallback_category_token(product_name) or "").strip()
    if not inferred or inferred == "Без категории" or inferred.casefold() == fallback.casefold():
        return ""
    return inferred


def _raw_supplier_inferred_category_for_product(product_name, known_categories):
    inferred = _canonical_ui_category_name(normalize_catalog_category_name(infer_category(product_name)))
    if not inferred or inferred == "Без категории":
        return ""
    fallback = str(_fallback_category_token(product_name) or "").strip()
    if inferred.casefold() == fallback.casefold() and inferred not in known_categories:
        return ""
    return inferred


def _file_cache_signature(path):
    try:
        stat = Path(path).stat()
        return (stat.st_mtime_ns, stat.st_size)
    except OSError:
        return (0, 0)


def _corrected_json_rows_cache_key(session_dir, cons_json_path, apply_visibility):
    state_keys = [CATEGORY_OVERRIDES_STATE, MANUAL_CATEGORY_OVERRIDES_STATE]
    if apply_visibility:
        state_keys.append(CATEGORY_VISIBILITY_STATE)
    return (
        str(Path(session_dir).resolve()),
        bool(apply_visibility),
        _file_cache_signature(cons_json_path),
        _category_state_signature(state_keys),
    )


def _get_corrected_json_rows_cache(key):
    with CORRECTED_JSON_ROWS_CACHE_LOCK:
        return CORRECTED_JSON_ROWS_CACHE.get(key)


def _set_corrected_json_rows_cache(key, rows):
    with CORRECTED_JSON_ROWS_CACHE_LOCK:
        CORRECTED_JSON_ROWS_CACHE[key] = rows
        while len(CORRECTED_JSON_ROWS_CACHE) > CORRECTED_JSON_ROWS_CACHE_LIMIT:
            oldest = next(iter(CORRECTED_JSON_ROWS_CACHE))
            CORRECTED_JSON_ROWS_CACHE.pop(oldest, None)


def _clear_corrected_json_rows_cache():
    with CORRECTED_JSON_ROWS_CACHE_LOCK:
        CORRECTED_JSON_ROWS_CACHE.clear()


def _normalize_visibility_map(visibility_map):
    hidden_categories = set()
    if not isinstance(visibility_map, dict):
        return {}
    if GLOBAL_VISIBILITY_KEY not in visibility_map:
        return {}
    categories = visibility_map.get(GLOBAL_VISIBILITY_KEY) or []
    existing = set()
    for category in categories:
        category_name = normalize_internal_category_name(category)
        if category_name:
            existing.add(category_name)
            existing.add(_canonical_ui_category_name(category_name))
            successors = CATEGORY_VISIBILITY_SUCCESSORS.get(category_name, set())
            existing.update(successors)
            existing.update(_canonical_ui_category_name(name) for name in successors)
    existing.discard("")
    hidden_categories.update(existing)
    if not hidden_categories:
        return {}
    return {GLOBAL_VISIBILITY_KEY: sorted(hidden_categories, key=_category_sort_key)}


def save_app_settings(settings):
    payload = _save_app_settings_base(settings)
    invalidate_onliner_b2b_token()
    return payload


def onliner_b2b_resolve_catalog_path_for_product(target_oid, product_name="", category_name="", force_refresh=False):
    return _b2b_resolve_catalog_path_for_product(
        target_oid,
        product_name=product_name,
        category_name=category_name,
        force_refresh=force_refresh,
        get_product_by_id=db_get_product_by_id,
        preferred_brand_token=_preferred_brand_token,
        extract_article=extract_article,
        priority_model_queries=_priority_model_queries,
        name_tokens=_name_tokens,
    )


def _fetch_onliner_market_stats_b2b(onliner_id, product_name="", category_name=""):
    return _b2b_fetch_market_stats(
        onliner_id,
        product_name=product_name,
        category_name=category_name,
        get_settings=get_onliner_b2b_settings,
        resolve_catalog_path=onliner_b2b_resolve_catalog_path_for_product,
        fetch_positions=onliner_b2b_fetch_product_positions_export,
    )


_ONLINER_MARKET_RUNTIME = None


def _get_onliner_market_runtime():
    global _ONLINER_MARKET_RUNTIME
    if _ONLINER_MARKET_RUNTIME is None:
        _ONLINER_MARKET_RUNTIME = OnlinerMarketRuntime(
            api_get=onliner_api_get,
            get_product_by_id=db_get_product_by_id,
            infer_category=infer_category,
            get_b2b_settings=get_onliner_b2b_settings,
            fetch_b2b_stats=_fetch_onliner_market_stats_b2b,
            read_consolidated_df=read_market_refresh_df,
            ensure_category_column=ensure_category_column,
            row_category=_precomputed_row_category,
            load_id_cache=load_id_cache,
            load_auto_refresh_settings=load_auto_refresh_settings,
            save_auto_refresh_settings=save_auto_refresh_settings,
            get_last_session_dir=lambda: LAST_ACTIVE_SESSION_DIR,
        )
    return _ONLINER_MARKET_RUNTIME


def onliner_b2b_search_candidates(local_name, category_name="", limit=30):
    return _b2b_search_candidates(
        local_name,
        category_name=category_name,
        limit=limit,
        preferred_brand_token=_preferred_brand_token,
        extract_article=extract_article,
        priority_model_queries=_priority_model_queries,
        name_tokens=_name_tokens,
        article_like_tokens=_article_like_tokens,
        strict_candidate_allowed=_strict_candidate_allowed,
        calc_name_match=calc_name_match,
        get_product_by_id=db_get_product_by_id,
        normalize_compact_name=_normalize_compact_name,
        upsert_product=db_upsert_product,
    )


def _fetch_api_source_worker(source_key, client_key):
    return _api_source_fetch_worker(
        source_key,
        client_key,
        upload_dir=UPLOAD_DIR,
        load_settings=load_app_settings,
        append_history=append_api_fetch_history,
    )


def get_no_id_category_rules(settings=None):
    settings = settings or load_app_settings()
    raw = str((((settings or {}).get("no_id_search") or {}).get("category_rules_text", "")) or "").strip()
    if not raw:
        return {}
    try:
        parsed = json.loads(raw)
        return parsed if isinstance(parsed, dict) else {}
    except Exception:
        return {}


# ============================================================
# ONLINER PRODUCTS DATABASE  (SQLite — единый источник истины)
# ============================================================

def _db_connection():
    return _onliner_db_connection()


def init_onliner_db():
    return _onliner_db_init()


def db_populate_from_df(df, source_label, skip_suppliers=None):
    return _onliner_db_populate_from_df(
        df,
        source_label,
        normalize_name_key=_normalize_name_key,
        skip_suppliers=skip_suppliers,
    )


def db_upsert_product(onliner_id, name, url, source="manual"):
    return _onliner_db_upsert_product(
        onliner_id,
        name,
        url,
        normalize_name_key=_normalize_name_key,
        source=source,
    )


def db_get_product_by_id(onliner_id):
    return _onliner_db_get_product_by_id(onliner_id)


def db_get_categories_by_ids(onliner_ids):
    return _onliner_db_get_categories_by_ids(onliner_ids)


def db_get_distinct_categories():
    return _onliner_db_get_distinct_categories()


def db_get_categories_by_exact_names(product_names):
    keys = sorted({_normalize_name_key(name) for name in product_names or [] if _normalize_name_key(name)})
    if not keys:
        return {}
    result = {}
    try:
        with _db_connection() as conn:
            for start in range(0, len(keys), 700):
                batch = keys[start:start + 700]
                placeholders = ",".join("?" for _ in batch)
                rows = conn.execute(
                    "SELECT ni.name_key, oc.category "
                    "FROM name_index ni JOIN onliner_catalog oc ON oc.onliner_id = ni.onliner_id "
                    f"WHERE ni.name_key IN ({placeholders}) "
                    "AND TRIM(COALESCE(oc.category, '')) <> ''",
                    batch,
                ).fetchall()
                for row in rows:
                    key = str(row["name_key"] or "").strip()
                    category = str(row["category"] or "").strip()
                    if key and category and key not in result:
                        result[key] = category
    except Exception as exc:
        APP_LOGGER.warning("DB exact-name category lookup failed: %s", exc)
    return result


def db_find_exact_id_for_name(product_name):
    return _onliner_db_find_exact_id(product_name, normalize_name_key=_normalize_name_key)


def db_find_id_for_name(product_name, threshold=0.75, allow_b2b=True):
    return _onliner_db_find_id(
        product_name,
        normalize_name_key=_normalize_name_key,
        raw_search_tokens=_raw_search_tokens,
        model_hint_tokens=_model_hint_tokens,
        preferred_brand_token=_preferred_brand_token,
        strict_candidate_allowed=_strict_candidate_allowed,
        calc_name_match=calc_name_match,
        b2b_search_candidates=onliner_b2b_search_candidates,
        threshold=threshold,
        allow_b2b=allow_b2b,
    )


def db_find_top_candidates(product_name, top_n=5, min_score=0.40, allow_b2b=True):
    return _onliner_db_find_top_candidates(
        product_name,
        raw_search_tokens=_raw_search_tokens,
        preferred_brand_token=_preferred_brand_token,
        strict_candidate_allowed=_strict_candidate_allowed,
        calc_name_match=calc_name_match,
        priority_model_queries=_priority_model_queries,
        b2b_search_candidates=onliner_b2b_search_candidates,
        top_n=top_n,
        min_score=min_score,
        allow_b2b=allow_b2b,
    )


def db_search_tgpc_pc_candidates(local_name, limit=12):
    return _onliner_db_search_tgpc_pc(
        local_name,
        find_exact=db_find_exact_id_for_name,
        tgpc_pc_code_queries=_tgpc_pc_code_queries,
        extract_tgpc_pc_code=_extract_tgpc_pc_code,
        is_tgpc_pc_name=_is_tgpc_pc_name,
        calc_name_match=calc_name_match,
        limit=limit,
    )


def db_search_iven_pc_candidates(local_name, limit=12):
    name = str(local_name or "").strip()
    if not name or not _is_iven_pc_name(name):
        return []
    limit = max(5, min(int(limit or 12), 80))
    local_code = _extract_iven_pc_code(name)
    candidates = []
    exact = db_find_exact_id_for_name(name)
    if isinstance(exact, dict):
        candidates.append({**exact, "source": exact.get("source") or "db_exact"})
    candidates.extend(_db_search_iven_pc_code_candidates(name, limit=limit))
    if not local_code:
        candidates.extend(db_find_top_candidates(name, top_n=limit, min_score=0.40, allow_b2b=False))

    out = []
    seen = set()
    for item in candidates:
        oid = normalize_onliner_id((item or {}).get("id", ""))
        if not oid or oid in seen:
            continue
        candidate_name = str((item or {}).get("name", "") or "").strip()
        candidate_url = str((item or {}).get("url", "") or "").strip()
        if not _is_iven_pc_candidate(candidate_name, candidate_url):
            continue
        if local_code:
            candidate_code = _extract_iven_pc_code(candidate_name)
            if not candidate_code or candidate_code != local_code:
                continue
        seen.add(oid)
        out.append({
            "id": oid,
            "name": candidate_name,
            "url": candidate_url,
            "score": float((item or {}).get("score", 0) or 0),
            "source": str((item or {}).get("source", "") or "db_iven_pc").strip(),
        })
        if len(out) >= limit:
            break
    out.sort(key=lambda item: float(item.get("score", 0.0) or 0.0), reverse=True)
    return out


def _db_search_iven_pc_code_candidates(local_name, limit=12):
    name = str(local_name or "").strip()
    queries = _iven_pc_code_queries(name)
    if not queries:
        return []
    local_code = _extract_iven_pc_code(name)
    local_series = _extract_iven_pc_series(name)
    pool = {}
    try:
        with _db_connection() as conn:
            fts_available = False
            if local_code:
                try:
                    rows = conn.execute(
                        "SELECT f.onliner_id, f.raw_name, oc.url "
                        "FROM name_index_fts f "
                        "LEFT JOIN onliner_catalog oc ON oc.onliner_id = f.onliner_id "
                        "WHERE name_index_fts MATCH ? LIMIT 180",
                        (f'"{local_code}"',),
                    ).fetchall()
                    fts_available = True
                except Exception:
                    rows = []
                for row in rows:
                    oid = normalize_onliner_id(row["onliner_id"])
                    raw_name = str(row["raw_name"] or "").strip()
                    url = str(row["url"] or "").strip()
                    if oid and raw_name:
                        pool[oid] = (raw_name, url)
            if not fts_available:
                for query in queries:
                    rows = conn.execute(
                        "SELECT ni.onliner_id, ni.raw_name, oc.url "
                        "FROM name_index ni JOIN onliner_catalog oc ON oc.onliner_id = ni.onliner_id "
                        "WHERE ni.raw_name LIKE ? LIMIT 180",
                        (f"%{query}%",),
                    ).fetchall()
                    for row in rows:
                        oid = normalize_onliner_id(row["onliner_id"])
                        raw_name = str(row["raw_name"] or "").strip()
                        url = str(row["url"] or "").strip()
                        if oid and raw_name:
                            pool[oid] = (raw_name, url)
    except Exception as exc:
        APP_LOGGER.warning("DB IVEN candidate search failed: %s", exc)
        return []

    items = []
    for oid, (candidate_name, candidate_url) in pool.items():
        if not _is_iven_pc_candidate(candidate_name, candidate_url):
            continue
        candidate_code = _extract_iven_pc_code(candidate_name)
        candidate_series = _extract_iven_pc_series(candidate_name)
        if local_code and candidate_code and local_code != candidate_code:
            continue
        match = calc_name_match(name, candidate_name)
        score = float(match.get("score", 0.0) or 0.0)
        if local_code and candidate_code == local_code:
            score = max(score, 0.98)
            if local_series and candidate_series and local_series == candidate_series:
                score = max(score, 0.995)
        if score < 0.40:
            continue
        items.append({
            "id": oid,
            "name": candidate_name,
            "url": candidate_url,
            "score": round(score, 3),
            "source": "db_iven_pc_code",
            "reason": str(match.get("reason", "") or ""),
        })
    items.sort(key=lambda item: float(item.get("score", 0.0) or 0.0), reverse=True)
    return items[:limit]


def db_stats():
    return _onliner_db_stats()


def _catalog_import_worker(filepath: str, file_ext: str, cleanup_file: bool = True):
    return _onliner_db_catalog_import_worker(
        filepath,
        file_ext,
        normalize_name_key=_normalize_name_key,
        cleanup_file=cleanup_file,
    )


def load_manual_id_bindings():
    return _manual_store_load_bindings()


def save_manual_id_bindings(bindings):
    _manual_store_save_bindings(bindings)


def load_review_queue():
    queue = _review_store_load_queue()
    migrate = globals().get("_migrate_review_queue_supplier_scope")
    if callable(migrate):
        queue, changed = migrate(queue)
        if changed:
            _review_store_save_queue(queue)
    return queue


def save_review_queue(queue):
    _review_store_save_queue(queue)


def load_id_change_journal():
    return _manual_store_load_journal()


def save_id_change_journal(rows):
    _manual_store_save_journal(rows)


def append_id_change_journal(entry):
    _manual_store_append_journal(entry)


def is_manually_confirmed_id(name, onliner_id, supplier_name=""):
    return _manual_store_is_confirmed_id(
        name,
        onliner_id,
        supplier_name=supplier_name,
        load_bindings=load_manual_id_bindings,
        normalize_name_key_func=_normalize_name_key,
    )


def _fetch_onliner_product_payload(onliner_id):
    return _get_onliner_market_runtime().fetch_product_payload(onliner_id)


def _fetch_onliner_market_stats_catalog_api(onliner_id):
    """Публичный catalog.api.onliner.by (fallback, если B2B выключен)."""
    return _get_onliner_market_runtime().fetch_catalog_market_stats(onliner_id)


def fetch_onliner_market_stats(onliner_id, product_name="", category_name=""):
    """Рыночные цены: сначала публичный catalog API (быстрее при массовом обновлении); при пустых данных — B2B price.api."""
    return _get_onliner_market_runtime().fetch_market_stats(onliner_id, product_name=product_name, category_name=category_name)


def get_onliner_market_stats_cached(onliner_id, cache=None):
    return _get_onliner_market_runtime().get_market_stats_cached(onliner_id, cache=cache)


def get_onliner_market_stats_bulk(onliner_ids, max_workers=22, id_hints=None):
    return _get_onliner_market_runtime().get_market_stats_bulk(onliner_ids, max_workers=max_workers, id_hints=id_hints)


def fetch_onliner_product_info(onliner_id, cache=None, force_refresh=False,
                               use_cache_on_error=True, product_name_hint=None):
    return _get_onliner_market_runtime().fetch_product_info(
        onliner_id,
        cache=cache,
        force_refresh=force_refresh,
        use_cache_on_error=use_cache_on_error,
        product_name_hint=product_name_hint,
    )


def search_onliner_product_by_name(local_name):
    """
    Fallback-поиск товара в Onliner API по названию локального товара.
    Возвращает лучший кандидат по score.
    """
    return _onliner_search_product_by_name(
        local_name,
        api_get=onliner_api_get,
        extract_article=extract_article,
        name_tokens=_name_tokens,
        calc_name_match=calc_name_match,
    )


def search_onliner_product_by_name_deep(local_name, category_name=""):
    """
    Более агрессивный поиск кандидата:
    - пробует обычный быстрый поиск,
    - затем расширенный поиск по нескольким запросам,
    - затем перепроверяет лучшие кандидаты по карточке товара.
    """
    return _onliner_search_product_by_name_deep(
        local_name,
        category_name=category_name,
        search_by_name=search_onliner_product_by_name,
        search_candidates=search_onliner_candidates,
        fetch_product_info=fetch_onliner_product_info,
        load_product_cache=load_onliner_product_cache,
        save_product_cache=save_onliner_product_cache,
        calc_name_match=calc_name_match,
        article_like_tokens=_article_like_tokens,
    )


def _category_path_hints(category_name):
    return _onliner_search_category_path_hints(category_name)


def _preferred_brand_token(text):
    """Return the first significant word that is NOT a category label and has no digits."""
    raw = str(text or "")
    skip = {
        "hdd", "ssd", "sata", "sataii", "sataiii", "usb", "usba", "usbc", "typea", "typec", "ddr3", "ddr4", "ddr5",
        "pc", "pcie", "nvme", "bulk", "tb", "gb", "mb", "mhz", "hz", "rpm", "mm", "cl", "pc25600",
        "dp", "dvi", "hdmi", "vga", "displayport", "minidp", "microdp",
        "microusb", "usbmicro", "jack", "male", "female",
        # Form-factor / descriptor words often following category names
        "atx", "matx", "eatx", "itx", "microatx", "miniitx", "midiтower", "minitower",
        # Russian prepositions/fillers common in product names
        "без", "для", "под", "про", "как", "из", "по", "на", "со", "за",
        # Common descriptor words
        "desktop", "mini", "midi", "tower", "slim", "ultra", "соединительный", "аудио", "стерео", "медь",
        # Russian category abbreviations that may follow form-factor
        "бп", "ибп", "мфу", "узу", "сзу",
        # "Без БП" = "without PSU" — skip both words
        "корпус", "корпуса", "кейс", "кейса",
        # Russian adjective qualifiers that prefix the product category
        "игровая", "игровой", "игровое", "беспроводная", "беспроводной",
        "офисная", "офисный", "офисное", "office",
        "проводная", "проводной", "механическая", "механический",
        "мембранная", "мембранный", "оптическая", "оптический",
        "лазерная", "лазерный", "ультратонкая", "ультратонкий",
        "портативная", "портативный", "внешний", "внешняя",
        "встроенный", "встроенная", "цветной", "цветная",
        "черно-белый", "черно-белая", "монохромный", "монохромная",
        # Russian category nouns (when they appear as first token after adjective)
        "мышь", "мышка", "клавиатура", "гарнитура", "наушники", "микрофон", "микрофоном", "монитор",
        "монитора", "мониторов", "жк", "кронштейн", "кронштейны",
        "принтер", "сканер", "колонки", "колонка", "акустика", "камера",
        "ноутбук", "ноутбука", "планшет", "смартфон", "телефон", "роутер", "коммутатор",
        "адаптер", "переходник", "кабель", "шнур", "разветвитель",
        "вентилятор", "кулер", "охладитель", "стабилизатор",
        "комплект", "комплекта", "набор", "набора", "вентиляторов",
        "сумка", "сумки", "сумку", "чехол", "чехла", "рюкзак",
        "накопитель", "накопители", "источник", "бесперебойного", "питания",
        "micro", "microsd", "micro-sd", "sd", "sdhc", "sdxc", "карта", "карты", "памяти",
    }
    # How many words to skip if name starts with a known multi-word category
    words_to_skip = 0
    stripped = raw.strip().lower()
    for n in (3, 2, 1):
        prefix_words = re.split(r"[\s,.(]+", stripped)[:n]
        prefix = " ".join(prefix_words)
        if prefix in _CATEGORY_LOOKUP:
            words_to_skip = n
            break

    idx = 0
    for token in re.findall(r"[A-Za-zА-Яа-я][A-Za-zА-Яа-я0-9.+_-]{1,}", raw):
        norm = _normalize_compact_name(token)
        if re.fullmatch(r"(?:dp|hdmi|dvi|usb|usba|usbc|typea|typec|jack|displayport)+", norm):
            idx += 1
            continue
        if not norm or norm in skip:
            idx += 1
            continue
        if any(ch.isdigit() for ch in norm):
            idx += 1
            continue
        if idx < words_to_skip:
            idx += 1
            continue
        return token
    return ""


def _normalized_brand_token(text):
    return _normalize_compact_name(_preferred_brand_token(text))


def _normalized_category_name(text):
    raw = str(text or "").strip()
    prefix_rules = [
        (r"^\s*внешн\w*\s+накопител", "Внешние накопители"),
        (r"^\s*(?:ибп\b|источник\s+бесперебойного\s+питания)", "ИБП"),
        (r"^\s*кронштейн", "Кронштейны"),
        (r"^\s*(?:web[- ]?cam|webcam|веб[- ]?камер)", "Веб-камеры"),
        (r"^\s*мфу\b", "МФУ"),
        (r"^\s*клавиатур", "Клавиатура"),
        (r"^\s*(?:мышь|мышка)", "Мышь"),
        (r"^\s*(?:наушник|гарнитур)", "Наушники"),
        (r"^\s*микрофон", "Микрофоны"),
        (r"^\s*(?:колонк|акустик)", "Акустика"),
        (r"^\s*usb\s*[23](?:[.,]\d+)?\b.*\b\d+\s*gb\b", "Накопители USB"),
    ]
    for pattern, category in prefix_rules:
        if re.search(pattern, raw, flags=re.IGNORECASE):
            return category
    category = str(normalize_catalog_category_name(infer_category(raw)) or "").strip()
    if category == "Без категории":
        return ""
    return category


def _normalized_categories_compatible(left, right):
    left = str(left or "").strip()
    right = str(right or "").strip()
    if not left or not right or left == right:
        return True
    compatible_groups = [
        {"Беспроводные адаптеры", "Сетевые адаптеры", "Сеть"},
        {"Кулер", "Кулеры", "Охлаждение"},
        {"Принтеры", "Принтер и МФУ", "МФУ"},
    ]
    if any(left in group and right in group for group in compatible_groups):
        return True
    peripheral_categories = {"Клавиатура", "Мышь", "Наушники", "Акустика", "Веб-камеры", "Микрофоны"}
    return bool(
        (left == "Периферия" and right in peripheral_categories)
        or (right == "Периферия" and left in peripheral_categories)
    )


def _strict_candidate_allowed(local_name, candidate_name):
    local = str(local_name or "").strip()
    candidate = str(candidate_name or "").strip()
    if not local or not candidate:
        return False, "empty"

    local_category = _normalized_category_name(local)
    candidate_category = _normalized_category_name(candidate)
    categories_compatible = _normalized_categories_compatible(local_category, candidate_category)
    local_prefix_category = _matching_extract_product_category(local)
    candidate_prefix_category = _matching_extract_product_category(candidate)
    if local_prefix_category and candidate_prefix_category and local_prefix_category != candidate_prefix_category:
        return False, "category"
    if local_prefix_category and not categories_compatible:
        return False, "category"

    comparison = calc_name_match(local, candidate)
    comparison_score = float(comparison.get("score", 0.0) or 0.0)
    comparison_reason = str(comparison.get("reason", "") or "")
    strong_identity_reasons = {
        "article",
        "article_like",
        "apple_article",
        "strict_article",
        "paren_model",
        "motherboard_model",
        "tgpc_code_exact",
    }
    if bool(comparison.get("match", False)) and (categories_compatible or not local_prefix_category) and (
        comparison_reason in strong_identity_reasons and comparison_score >= 0.74
    ):
        return True, "strong_identity"

    if not categories_compatible:
        return False, "category"

    local_brand = _normalized_brand_token(local)
    candidate_brand = _normalized_brand_token(candidate)
    if local_brand:
        if not candidate_brand:
            return False, "brand_missing"
        if local_brand != candidate_brand:
            local_compact = _normalize_compact_name(local)
            candidate_compact = _normalize_compact_name(candidate)
            if candidate_brand not in local_compact and local_brand not in candidate_compact:
                return False, "brand"

    return True, "ok"


def _priority_model_queries(text):
    raw = str(text or "").strip()
    if not raw:
        return []

    seen = set()
    out = []
    brand = _preferred_brand_token(raw)

    def _add(query):
        q = str(query or "").strip()
        key = q.lower()
        if not q or key in seen:
            return
        seen.add(key)
        out.append(q)

    def _push_token(token):
        token = str(token or "").strip()
        norm = _normalize_compact_name(token)
        token_lower = token.lower()
        if len(norm) < 5:
            return
        if not any(ch.isdigit() for ch in norm):
            return
        if not any(ch.isalpha() for ch in norm):
            return
        if re.match(r"^\d{2,4}x\d{2,4}$", token_lower):
            return
        _add(token)
        if brand and norm != _normalize_compact_name(brand):
            _add(f"{brand} {token}")

    for chunk in _paren_chunks(raw):
        for token in extract_article_candidates(chunk):
            _push_token(token)
        for token in re.findall(r"[A-Za-z0-9][A-Za-z0-9._\-/]{4,}", chunk):
            _push_token(token)

    for token in extract_article_candidates(raw):
        _push_token(token)

    for token in re.findall(r"[A-Za-z0-9][A-Za-z0-9._\-/]{4,}", raw):
        _push_token(token)

    return out


def _tgpc_pc_code_queries(text):
    raw = str(text or "").strip()
    if not raw:
        return []

    queries = []
    seen = set()

    def _add(value):
        q = str(value or "").strip()
        key = q.lower()
        if not q or key in seen:
            return
        seen.add(key)
        queries.append(q)

    # TGPC ПЭВМ: "ПЭВМ TGPC Action 5 81872 A-X Ryzen 5 ..."
    # Приоритетный ключ для Onliner: "81872 A-X"
    m = re.search(r"\b(\d{4,6})\s+([A-ZА-Я]-X)\b", raw, flags=re.IGNORECASE)
    if m:
        code = m.group(1).strip()
        suffix = m.group(2).upper().replace("А", "A")
        _add(f"{code} {suffix}")
        _add(f"{code}{suffix.replace('-', '')}")

    # Иногда формат бывает слитный: 81872A-X
    m2 = re.search(r"\b(\d{4,6})([A-ZА-Я]-X)\b", raw, flags=re.IGNORECASE)
    if m2:
        code = m2.group(1).strip()
        suffix = m2.group(2).upper().replace("А", "A")
        _add(f"{code} {suffix}")
        _add(f"{code}{suffix.replace('-', '')}")

    return queries


def _extract_iven_pc_code(text):
    raw = str(text or "")
    series_match = re.search(
        r"\b(?:iven|ивен)(?:\s+by)?\s+"
        r"(?:gaming|office|home|pro|ultra|superpower|gamebasic)\b.*?\b(\d{5,6})\b",
        raw,
        flags=re.IGNORECASE,
    )
    if series_match:
        return series_match.group(1)
    candidates = re.findall(r"\b(\d{5,6})\b", raw)
    if not candidates:
        return ""
    for code in candidates:
        if code.startswith(("17", "18", "19", "20", "21")):
            return code
    return candidates[0]


def _extract_iven_pc_series(text):
    raw = str(text or "").strip().lower()
    for series in ("office", "gaming", "home", "pro", "ultra", "superpower", "gamebasic"):
        if re.search(rf"\b{series}\b", raw, flags=re.IGNORECASE):
            return series
    return ""


def _iven_pc_code_queries(text):
    raw = str(text or "").strip()
    code = _extract_iven_pc_code(raw)
    if not code:
        return []
    series = _extract_iven_pc_series(raw)
    queries = []
    seen = set()

    def _add(value):
        q = str(value or "").strip()
        key = q.lower()
        if not q or key in seen:
            return
        seen.add(key)
        queries.append(q)

    if series:
        _add(f"Iven {series.title()} {code}")
        _add(f"IVEN BY {series.title()} {code}")
        _add(f"{series.title()} {code}")
    _add(code)
    return queries


def _is_tgpc_pc_name(text):
    raw = str(text or "").strip().lower()
    if "tgpc" not in raw:
        return False
    if _tgpc_pc_code_queries(raw):
        return True
    return any(token in raw for token in ["action", "mesh", "osprey", "xtreme", "valise", "gaming"])


def _is_iven_pc_name(text):
    raw = str(text or "").strip().lower()
    if not raw:
        return False
    if raw.startswith(("компьютер ", "системный блок ", "пэвм ")):
        return "iven" in raw or "ивен" in raw
    if "iven" not in raw and "ивен" not in raw:
        return False
    return bool(re.search(r"\b(?:gaming|office|home|pro|ultra|superpower|gamebasic)\b", raw, flags=re.IGNORECASE))


def _is_iven_pc_candidate(name, url=""):
    raw = str(name or "").strip().lower()
    link = str(url or "").strip().lower()
    if _is_iven_pc_name(raw):
        return True
    if "iven" in raw and ("компьютер" in raw or "системный блок" in raw or "пэвм" in raw):
        return True
    return bool(link and any(part in link for part in ("/desktop/", "/desktoppc/", "/computer/", "/monoblock/", "/nettop/")))


def search_onliner_candidates(local_name, category_name="", query="", limit=80, max_queries=4, timeout_sec=6):
    return _onliner_search_candidates(
        local_name,
        category_name=category_name,
        query=query,
        limit=limit,
        max_queries=max_queries,
        timeout_sec=timeout_sec,
        load_settings=load_app_settings,
        get_category_rules=get_no_id_category_rules,
        coerce_bool=_coerce_bool,
        api_get=onliner_api_get,
        b2b_search_candidates=onliner_b2b_search_candidates,
        extract_article=extract_article,
        name_tokens=_name_tokens,
        preferred_brand_token=_preferred_brand_token,
        normalize_compact_name=_normalize_compact_name,
        priority_model_queries=_priority_model_queries,
        tgpc_pc_code_queries=_tgpc_pc_code_queries,
        is_tgpc_pc_name=_is_tgpc_pc_name,
        extract_tgpc_pc_code=_extract_tgpc_pc_code,
        model_hint_tokens=_model_hint_tokens,
        paren_chunks=_paren_chunks,
        article_like_tokens=_article_like_tokens,
        token_family_match=_token_family_match,
        strict_candidate_allowed=_strict_candidate_allowed,
        calc_name_match=calc_name_match,
        query_cache=ID_REPLACE_QUERY_CACHE,
        query_cache_lock=ID_REPLACE_QUERY_CACHE_LOCK,
        query_cache_ttl=ID_REPLACE_QUERY_CACHE_TTL,
        query_cache_version=ID_REPLACE_QUERY_CACHE_VERSION,
    )


def _name_tokens(text):
    return _matching_name_tokens(text)


def _normalize_compact_name(text):
    return _matching_normalize_compact_name(text)


def _normalize_match_text(text):
    return _matching_normalize_match_text(text)


def _paren_chunks(text):
    return _matching_paren_chunks(text)


def _is_color_only_chunk(text):
    return _matching_is_color_only_chunk(text)


def _model_hint_tokens(text):
    return _matching_model_hint_tokens(
        text,
        tgpc_pc_code_queries=_tgpc_pc_code_queries,
        extract_article_candidates=extract_article_candidates,
    )


def _token_family_match(left_tokens, right_tokens):
    return _matching_token_family_match(left_tokens, right_tokens)


def _capacity_tokens(text):
    return _matching_capacity_tokens(text)


def _important_name_tokens(text):
    return _matching_important_name_tokens(text)


def _ordered_token_hits(left_tokens, right_tokens):
    return _matching_ordered_token_hits(left_tokens, right_tokens)


def _color_tokens(text):
    return _matching_color_tokens(text)


def _extract_product_category(name: str):
    return _matching_extract_product_category(name)


def _is_spec_code(norm_upper: str) -> bool:
    return _matching_is_spec_code(norm_upper)


def _raw_paren_article_tokens(text):
    return _matching_raw_paren_article_tokens(text)


def _raw_search_tokens(text):
    return _matching_raw_search_tokens(text)


def _article_like_tokens(text):
    return _matching_article_like_tokens(
        text,
        tgpc_pc_code_queries=_tgpc_pc_code_queries,
        extract_article_candidates=extract_article_candidates,
    )


def _extract_tgpc_pc_code(text):
    return _matching_extract_tgpc_pc_code(text)


def _extract_gpu_model(text):
    return _matching_extract_gpu_model(text)


def calc_name_match(local_name, onliner_name):
    return _matching_calc_name_match(
        local_name,
        onliner_name,
        extract_article=extract_article,
        preferred_brand_token=_preferred_brand_token,
        tgpc_pc_code_queries=_tgpc_pc_code_queries,
        extract_article_candidates=extract_article_candidates,
    )


def _harden_base_verify_result(oid, local_name, verify_result):
    return _matching_harden_base_verify_result(
        local_name,
        verify_result,
        lookup_catalog_match_details=lookup_catalog_match_details,
        calc_match=calc_name_match,
        article_tokens=_article_like_tokens,
    )


def _openai_autosort_predict_category(product_name, categories, local_hint=""):
    return _autosort_predict_openai_category(
        product_name,
        categories,
        local_hint=local_hint,
        api_key=OPENAI_API_KEY,
        model=OPENAI_AUTOSORT_MODEL,
        timeout_sec=OPENAI_AUTOSORT_TIMEOUT_SEC,
        cache=AI_CATEGORY_CACHE,
        cache_lock=AI_CATEGORY_CACHE_LOCK,
        requests_post=requests.post,
    )


def reconcile_ids_from_catalog(df):
    return _catalog_reconcile_ids(
        df,
        lookup_id_from_catalog_sheet=lookup_id_from_catalog_sheet,
        normalize_onliner_id=normalize_onliner_id,
    )


def enforce_catalog_consistency(df, session_dir=None):
    try:
        return _catalog_enforce_consistency(
            df,
            session_dir=session_dir,
            lookup_id_from_catalog_sheet=lookup_id_from_catalog_sheet,
            normalize_onliner_id=normalize_onliner_id,
            get_article_from_name=get_article_from_name,
            save_summary=save_dict,
        )
    except Exception as e:
        APP_LOGGER.warning("ID quality report save failed: %s", e)
        return _catalog_enforce_consistency(
            df,
            session_dir=None,
            lookup_id_from_catalog_sheet=lookup_id_from_catalog_sheet,
            normalize_onliner_id=normalize_onliner_id,
            get_article_from_name=get_article_from_name,
        )


def get_effective_category(row, overrides=None):
    if overrides is None:
        overrides = load_category_overrides()
    return _category_get_effective(
        row,
        overrides=overrides,
        build_item_category_keys=build_item_category_keys,
        infer_category=infer_category,
    )


def row_category(row, overrides=None):
    if overrides is None:
        overrides = load_category_overrides()
    return _category_row_category(
        row,
        overrides=overrides,
        build_item_category_keys=build_item_category_keys,
        infer_category=infer_category,
    )


def ensure_category_column(df, overrides=None):
    if overrides is None:
        overrides = load_category_overrides()
    if df is None or df.empty or "OnlinerID" not in df.columns:
        return _category_ensure_column(
            df,
            overrides=overrides,
            build_item_category_keys=build_item_category_keys,
            infer_category=infer_category,
        )

    ids = [normalize_onliner_id(value) for value in df["OnlinerID"].tolist()]
    catalog_categories = db_get_categories_by_ids(ids)
    if not catalog_categories:
        return _category_ensure_column(
            df,
            overrides=overrides,
            build_item_category_keys=build_item_category_keys,
            infer_category=infer_category,
        )

    out = df.copy()
    if "Категория" not in out.columns:
        out["Категория"] = ""
    unresolved_indices = []
    has_name = "Название" in out.columns
    for index, oid in zip(out.index, ids):
        raw_category = str(catalog_categories.get(oid, "") or "").strip()
        product_name = out.at[index, "Название"] if has_name else ""
        category = _native_catalog_category_for_product(raw_category, product_name)
        if category:
            out.at[index, "Категория"] = category
        else:
            unresolved_indices.append(index)

    if unresolved_indices:
        unresolved = _category_ensure_column(
            out.loc[unresolved_indices].copy(),
            overrides=overrides,
            build_item_category_keys=build_item_category_keys,
            infer_category=infer_category,
        )
        out.loc[unresolved_indices, "Категория"] = unresolved["Категория"]
    return out


def apply_onliner_catalog_categories(df):
    """Use the Onliner catalog category whenever a row already has a valid OnlinerID."""
    if df is None or df.empty or "OnlinerID" not in df.columns:
        return df
    ids = [normalize_onliner_id(value) for value in df["OnlinerID"].tolist()]
    catalog_categories = db_get_categories_by_ids(ids)
    if not catalog_categories:
        return df
    df = df.copy()
    for index, oid in zip(df.index, ids):
        raw_category = str(catalog_categories.get(oid, "") or "").strip()
        category = _native_catalog_category_for_product(raw_category, df.at[index, "Название"] if "Название" in df.columns else "")
        if category:
            df.at[index, "Категория"] = category
    return df


def load_visibility_map(session_dir):
    with CATEGORY_VISIBILITY_LOCK:
        visibility_map = _category_load_visibility(
            session_dir,
            load_visibility=lambda: _category_state_load(
                CATEGORY_VISIBILITY_STATE,
                CATEGORY_VISIBILITY_FILE,
                sqlite_primary=Path(CATEGORY_VISIBILITY_FILE) == DEFAULT_CATEGORY_VISIBILITY_FILE,
            ),
        )
        return _normalize_visibility_map(visibility_map)


def save_visibility_map(session_dir, visibility_map):
    with CATEGORY_VISIBILITY_LOCK:
        return _category_save_visibility(
            session_dir,
            _normalize_visibility_map(visibility_map),
            save_visibility=lambda payload: _category_state_save(
                _normalize_visibility_map(payload),
                CATEGORY_VISIBILITY_STATE,
                CATEGORY_VISIBILITY_FILE,
                sqlite_primary=Path(CATEGORY_VISIBILITY_FILE) == DEFAULT_CATEGORY_VISIBILITY_FILE,
            ),
        )


def apply_saved_markups_to_df(df):
    return _category_apply_saved_markups(
        df,
        load_category_markups=load_category_markups,
        load_category_overrides=load_category_overrides,
        get_category_markup_config=get_category_markup_config,
        calc_rrc_and_no_discount=calc_rrc_and_no_discount,
        normalize_onliner_id=normalize_onliner_id,
        get_onliner_market_stats_from_cache_only=get_onliner_market_stats_from_cache_only,
        build_item_category_keys=build_item_category_keys,
        infer_category=infer_category,
    )


def apply_visibility_filter(df, session_dir):
    return _category_apply_visibility_filter(
        df,
        session_dir,
        load_visibility_map_func=load_visibility_map,
        load_category_overrides=load_category_overrides,
        build_item_category_keys=build_item_category_keys,
        infer_category=infer_category,
        normalize_supplier=_canonical_supplier_name,
        normalize_category=_canonical_ui_category_name,
    )


def apply_quality_visibility_filter(df, session_dir):
    return _quality_apply_visibility_filter(
        df,
        session_dir,
        apply_visibility_filter=apply_visibility_filter,
        load_visibility_map=load_visibility_map,
        normalize_category=_canonical_ui_category_name,
    )


def _safe_json_value(value):
    return _consolidated_safe_json_value(value)


def _delivery_days_from_row(row):
    return _consolidated_delivery_days_from_row(row)


def write_consolidated_json(df, json_path):
    return _consolidated_write_json(df, json_path)


def read_consolidated_json_rows(json_path):
    return _consolidated_read_json_rows(json_path)


def read_consolidated_df(session_dir):
    return _consolidated_read_df(session_dir)


def read_market_refresh_df(session_dir):
    """Read the freshest consolidated data for market refresh without slow XLSX parsing."""
    try:
        df = _consolidated_json_df(session_dir, apply_visibility=True)
        if df is not None:
            return df
    except Exception:
        pass
    return read_consolidated_df(session_dir)


def read_consolidated_json_fast_df(session_dir):
    """Read the current consolidated session from JSON; fall back to XLSX only if JSON is absent."""
    try:
        df = _consolidated_json_df(session_dir, apply_visibility=False)
        if df is not None:
            return df
    except Exception:
        pass
    return read_consolidated_df(session_dir)


def read_consolidated_export_df(session_dir):
    """Read consolidated data and overlay durable manual ID bindings before export."""
    df = _apply_manual_bindings_to_consolidated_df(read_consolidated_json_fast_df(session_dir))
    return apply_saved_markups_to_df(df)


def _has_consolidated_session_file(session_dir):
    if not session_dir:
        return False
    session_path = Path(session_dir)
    return (session_path / "consolidated.json").exists() or (session_path / "consolidated_price.xlsx").exists()


def write_consolidated_df(session_dir, df):
    return _consolidated_write_df(session_dir, df)


def write_consolidated_df_background(session_dir, df, *, label="consolidated"):
    try:
        return BACKGROUND_XLSX_WORKER.enqueue(session_dir, df, label=label)
    except Exception as exc:
        APP_LOGGER.exception("background XLSX queue failed label=%s", label)
        return {"state": "error", "running": False, "message": str(exc)}


@app.get("/api/background-xlsx-status")
def api_background_xlsx_status():
    session_dir = get_active_session_dir()
    return jsonify(BACKGROUND_XLSX_WORKER.status(session_dir))


@app.get("/api/worker-status")
def api_worker_status():
    if DURABLE_JOB_QUEUE is None:
        return jsonify({
            "mode": "inline",
            "status": "ok",
            "queue": {},
            "active_workers": 0,
        })
    health = DURABLE_JOB_QUEUE.worker_health()
    return jsonify({
        "mode": "external",
        **health,
        "queue": DURABLE_JOB_QUEUE.counts(),
    }), (200 if health["status"] == "ok" else 503)


def _create_session_dir():
    return _upload_sessions_create_dir(UPLOAD_DIR)


def _finalize_processed_session(session_id, session_dir, output_path):
    global LAST_ACTIVE_SESSION_DIR
    session["session_id"] = session_id
    session.pop("output_path", None)
    session.pop("session_dir", None)
    LAST_ACTIVE_SESSION_DIR = str(session_dir)


def _process_supplier_files(file_entries, session_id=None, session_dir=None):
    return _processing_process_supplier_files(
        file_entries,
        session_id=session_id,
        session_dir=session_dir,
        create_session_dir=_create_session_dir,
        load_app_settings=load_app_settings,
        parse_generic_excel=parse_generic_excel,
        consolidate_simple=consolidate_simple,
        normalize_consolidated_columns=normalize_consolidated_columns,
        ensure_category_column=ensure_category_column,
        apply_saved_markups_to_df=apply_saved_markups_to_df,
        load_manual_id_bindings=load_manual_id_bindings,
        expand_iven_pc_manual_aliases=_expand_iven_pc_manual_aliases,
        save_manual_id_bindings=save_manual_id_bindings,
        load_id_cache=load_id_cache,
        sanitize_id_cache=_sanitize_id_cache,
        save_id_cache=save_id_cache,
        build_id_fanout_map=build_id_fanout_map,
        normalize_name_key=_normalize_name_key,
        normalize_onliner_id=normalize_onliner_id,
        is_iven_pc_name=_is_iven_pc_name,
        iven_pc_onliner_id_mismatch_known=_iven_pc_onliner_id_mismatch_known,
        allow_manual_binding_for_supplier=_allow_manual_binding_for_supplier,
        lookup_manual_binding_for_name=_lookup_manual_binding_for_name,
        id_cache_keys_for_iven_pc_name=_id_cache_keys_for_iven_pc_name,
        get_id_cache_key_for_name=_get_id_cache_key_for_name,
        is_trusted_cached_id=is_trusted_cached_id,
        iven_pc_onliner_id_matches_name=_iven_pc_onliner_id_matches_name,
        clear_duplicate_onliner_ids_for_suppliers=_clear_duplicate_onliner_ids_for_suppliers,
        write_consolidated_df=write_consolidated_df,
        write_consolidated_json=write_consolidated_json,
        save_session_supplier_diff=save_session_supplier_diff,
        count_rows_without_onliner_id=_count_rows_without_onliner_id,
        count_rows_with_duplicate_onliner_id=_count_rows_with_duplicate_onliner_id,
        coerce_bool=_coerce_bool,
        maybe_cleanup_old_uploads=_maybe_cleanup_old_uploads,
        last_active_session_dir=LAST_ACTIVE_SESSION_DIR,
    )


def _infer_supplier_from_filename(filename, app_settings=None):
    return _processing_infer_supplier_from_filename(
        filename,
        app_settings=app_settings,
        load_app_settings=load_app_settings,
    )


def upload():
    wants_json = request.headers.get("X-Requested-With") == "XMLHttpRequest"
    files = request.files.getlist("files")
    APP_LOGGER.info(
        "upload received file_count=%s",
        len([item for item in files if item and item.filename]),
    )
    if not files or all(not f.filename for f in files):
        if wants_json:
            return jsonify({"status": "error", "message": "Не загружено ни одного файла"}), 400
        return redirect(url_for("main_api.index", error="Не загружено ни одного файла"))

    session_id, session_dir = _create_session_dir()
    app_settings = load_app_settings()
    try:
        file_entries = _upload_files_build_entries(
            files,
            request.form,
            session_dir,
            app_settings,
            _infer_supplier_from_filename,
        )
    except ValueError as exc:
        shutil.rmtree(session_dir, ignore_errors=True)
        message = str(exc)[:180]
        if wants_json:
            return jsonify({"status": "error", "message": message}), 400
        return redirect(url_for("main_api.index", error=message))

    try:
        result = _process_supplier_files(file_entries, session_id=session_id, session_dir=session_dir)
    except Exception as _upload_err:
        APP_LOGGER.exception("upload processing failed")
        message = "Не удалось обработать файлы: " + str(_upload_err)[:120]
        if wants_json:
            return jsonify({"status": "error", "message": message}), 500
        return redirect(url_for("main_api.index", error=message))

    _finalize_processed_session(result["session_id"], result["session_dir"], result["output_path"])
    redirect_url = url_for("main_api.result_page", sid=result["session_id"])
    if wants_json:
        return jsonify({"status": "ok", "redirect_url": redirect_url})
    return redirect(redirect_url)


def _correct_consolidated_json_rows(session_dir, *, apply_visibility=True):
    cons_json_path = Path(session_dir) / "consolidated.json"
    if not cons_json_path.exists():
        return None
    cache_key = _corrected_json_rows_cache_key(session_dir, cons_json_path, apply_visibility)
    cached_rows = _get_corrected_json_rows_cache(cache_key)
    if cached_rows is not None:
        return cached_rows
    if apply_visibility:
        corrected_data = _correct_consolidated_json_rows(session_dir, apply_visibility=False)
        if corrected_data is None:
            return None
        visibility_map = load_visibility_map(session_dir)
        if visibility_map:
            hidden_categories = {
                _canonical_ui_category_name(category)
                for categories in visibility_map.values()
                for category in categories
                if str(category or "").strip()
            }
            corrected_data = [
                row for row in corrected_data
                if _canonical_ui_category_name(row[9]) not in hidden_categories
            ]
        _set_corrected_json_rows_cache(cache_key, corrected_data)
        return corrected_data
    cons_data = read_consolidated_json_rows(cons_json_path)
    if not cons_data or not all(len(row) >= 10 for row in cons_data):
        return None

    catalog_categories = db_get_categories_by_ids([row[0] for row in cons_data])
    exact_name_categories = db_get_categories_by_exact_names([row[1] for row in cons_data])
    overrides = load_category_overrides()
    explicit_overrides = load_manual_category_overrides()
    known_raw_infer_categories = _supplier_visibility_known_categories()
    corrected_data = []
    for row in cons_data:
        current_category = normalize_internal_category_name(row[9])
        row_category_name = current_category
        row_item = {"Название": row[1], "Поставщик": row[3], "Категория": current_category}
        explicit_category = _repair_saved_category_for_product(
            _category_override_for_row(row_item, explicit_overrides),
            row[1],
        )
        manual_category = _repair_saved_category_for_product(
            _category_override_for_row(row_item, overrides),
            row[1],
        )
        if _looks_like_raw_supplier_category(explicit_category):
            explicit_category = ""
        if _looks_like_raw_supplier_category(manual_category):
            manual_category = ""
        onliner_id = normalize_onliner_id(row[0])
        catalog_category = _native_catalog_category_for_product(catalog_categories.get(onliner_id, ""), row[1])
        exact_name_category = _native_catalog_category_for_product(exact_name_categories.get(_normalize_name_key(row[1]), ""), row[1])
        raw_inferred_category = ""
        if _looks_like_raw_supplier_category(current_category):
            raw_inferred_category = _raw_supplier_inferred_category_for_product(row[1], known_raw_infer_categories)
        if onliner_id and catalog_category:
            row_category_name = catalog_category
        elif onliner_id and exact_name_category:
            row_category_name = exact_name_category
        elif not onliner_id and exact_name_category:
            row_category_name = exact_name_category
        elif explicit_category:
            row_category_name = explicit_category
        elif onliner_id:
            row_category_name = _strong_inferred_category_for_product(row[1]) or _sorting_review_category(current_category)
        elif raw_inferred_category:
            row_category_name = raw_inferred_category
        elif manual_category:
            row_category_name = manual_category
        elif _json_row_needs_category_repair(row[1], row[9], current_category):
            category = _category_row_category(
                {"Название": row[1], "Поставщик": row[3], "Категория": current_category},
                overrides={},
                build_item_category_keys=build_item_category_keys,
                infer_category=infer_category,
            )
            row_category_name = normalize_internal_category_name(category)
        row_category_name = _canonical_ui_category_name(row_category_name)
        if _looks_like_raw_supplier_category(row_category_name):
            row_category_name = _sorting_review_category(row_category_name)
        if row_category_name != row[9]:
            row = list(row)
            row[9] = row_category_name
        corrected_data.append(row)

    _set_corrected_json_rows_cache(cache_key, corrected_data)
    return corrected_data


SORTING_REPARSE_URL = os.getenv(
    "SORTING_REPARSE_URL",
    "http://127.0.0.1:5055/api/price-mixer",
).rstrip("/")
SORTING_REPARSE_MONITOR_LOCK = threading.Lock()
SORTING_REPARSE_START_LOCK = threading.Lock()
sorting_reparse_monitor_running = False


def _sorting_reparse_service_healthy(timeout=1.0):
    try:
        response = requests.get(f"{SORTING_REPARSE_URL}/status", timeout=timeout)
        return bool(response.ok)
    except requests.RequestException:
        return False


def _sorting_reparse_launch_spec():
    configured_dir = os.getenv("ONLINER_PARSER_DIR", "").strip()
    candidate_dirs = []
    if configured_dir:
        candidate_dirs.append(Path(configured_dir).expanduser())
    candidate_dirs.extend([
        Path(__file__).resolve().parent.parent / "onliner-parser",
        Path("/opt/onliner-parser"),
    ])
    for parser_dir in candidate_dirs:
        script_path = parser_dir / "ui_server.py"
        if not script_path.is_file():
            continue
        configured_python = os.getenv("ONLINER_PARSER_PYTHON", "").strip()
        python_candidates = [
            Path(configured_python).expanduser() if configured_python else None,
            parser_dir / ".venv" / "bin" / "python",
            parser_dir / ".venv" / "bin" / "python3",
            Path(sys.executable),
        ]
        python_path = next((path for path in python_candidates if path and path.is_file()), None)
        if python_path:
            return [str(python_path), str(script_path)], parser_dir, parser_dir / "parser_stdout.log"
    raise RuntimeError(
        "не найден onliner-parser/ui_server.py; укажи ONLINER_PARSER_DIR в .env"
    )


def _ensure_sorting_reparse_service(start_timeout=8.0):
    if _sorting_reparse_service_healthy():
        return
    with SORTING_REPARSE_START_LOCK:
        if _sorting_reparse_service_healthy():
            return
        command, parser_dir, log_path = _sorting_reparse_launch_spec()
        with log_path.open("a", encoding="utf-8") as log_file:
            subprocess.Popen(
                command,
                cwd=str(parser_dir),
                stdin=subprocess.DEVNULL,
                stdout=log_file,
                stderr=subprocess.STDOUT,
                start_new_session=True,
                close_fds=True,
            )
        deadline = time.monotonic() + max(float(start_timeout), 1.0)
        while time.monotonic() < deadline:
            if _sorting_reparse_service_healthy(timeout=0.8):
                return
            time.sleep(0.2)
    raise RuntimeError("парсер не запустился за отведённое время; проверь parser_stdout.log")


def _sorting_reparse_items():
    session_dir = get_active_session_dir()
    rows = _correct_consolidated_json_rows(session_dir) if session_dir else None
    items = []
    seen = set()
    for row in rows or []:
        oid = normalize_onliner_id(row[0])
        category = str(row[9] or "").strip()
        if not oid or oid in seen or not _is_sorting_review_category(category):
            continue
        seen.add(oid)
        items.append({
            "onliner_id": oid,
            "name": str(row[1] or "").strip(),
            "parent_category": category[len(SORTING_REVIEW_PREFIX):].strip(),
        })
    return items


def _all_onliner_reparse_items():
    session_dir = get_active_session_dir()
    rows = _correct_consolidated_json_rows(session_dir, apply_visibility=False) if session_dir else None
    known_categories = db_get_categories_by_ids([row[0] for row in rows or []])
    items = []
    seen = set()
    for row in rows or []:
        oid = normalize_onliner_id(row[0])
        if not oid or oid in seen:
            continue
        current_category = str(row[9] or "").strip()
        known_category = _native_catalog_category_for_product(known_categories.get(oid, ""), row[1])
        if known_category and not _is_sorting_review_category(current_category):
            seen.add(oid)
            continue
        seen.add(oid)
        items.append({
            "onliner_id": oid,
            "name": str(row[1] or "").strip(),
            "parent_category": current_category,
            "strict_api": True,
        })
    return items


def _write_sorting_reparse_results_to_db(results):
    written = _onliner_db_update_categories(results) if results else 0
    if written:
        _clear_corrected_json_rows_cache()
    return written


def _start_sorting_reparse_monitor():
    global sorting_reparse_monitor_running
    with SORTING_REPARSE_MONITOR_LOCK:
        if sorting_reparse_monitor_running:
            return
        sorting_reparse_monitor_running = True

    def worker():
        global sorting_reparse_monitor_running
        try:
            while True:
                try:
                    response = requests.get(f"{SORTING_REPARSE_URL}/status", timeout=15)
                    payload = _sorting_reparse_json_payload(response)
                    results = payload.get("results") or []
                    _write_sorting_reparse_results_to_db(results)
                    if not payload.get("is_running"):
                        break
                except Exception as exc:
                    APP_LOGGER.warning("sorting reparse monitor failed: %s", exc)
                time.sleep(2)
        finally:
            with SORTING_REPARSE_MONITOR_LOCK:
                sorting_reparse_monitor_running = False

    threading.Thread(target=worker, daemon=True).start()


@app.post("/api/sorting-reparse/run")
def api_sorting_reparse_run():
    items = _sorting_reparse_items()
    if not items:
        return jsonify({"ok": False, "error": "Очередь «Требует сортировки» пуста."}), 400
    try:
        _ensure_sorting_reparse_service()
        response = requests.post(f"{SORTING_REPARSE_URL}/run", json={"items": items}, timeout=15)
        payload = _sorting_reparse_json_payload(response)
        if response.ok and payload.get("ok"):
            _start_sorting_reparse_monitor()
        return jsonify(payload), response.status_code
    except Exception as exc:
        return jsonify({"ok": False, "error": _sorting_reparse_error_message(exc)}), 502


@app.post("/api/sorting-reparse/run-all")
def api_sorting_reparse_run_all():
    items = _all_onliner_reparse_items()
    if not items:
        return jsonify({"ok": False, "error": "В текущем прайсе нет товаров с OnlinerID."}), 400
    try:
        _ensure_sorting_reparse_service()
        response = requests.post(f"{SORTING_REPARSE_URL}/run", json={"items": items}, timeout=15)
        payload = _sorting_reparse_json_payload(response)
        if response.ok and payload.get("ok"):
            _start_sorting_reparse_monitor()
        return jsonify(payload), response.status_code
    except Exception as exc:
        return jsonify({"ok": False, "error": _sorting_reparse_error_message(exc)}), 502


@app.get("/api/sorting-reparse/status")
def api_sorting_reparse_status():
    try:
        response = requests.get(f"{SORTING_REPARSE_URL}/status", timeout=15)
        payload = _sorting_reparse_json_payload(response)
    except Exception as exc:
        return jsonify({"ok": False, "error": _sorting_reparse_error_message(exc)}), 502
    results = payload.get("results") or []
    payload["written_to_db"] = _write_sorting_reparse_results_to_db(results)
    payload["ok"] = True
    return jsonify(payload), response.status_code


def _json_row_needs_category_repair(name, raw_category, current_category):
    if not current_category:
        return True
    raw = str(raw_category or "").strip()
    if normalize_internal_category_name(raw) != raw:
        return True
    current = str(current_category or "").strip()
    text = str(name or "").strip().lower()
    text_without_prefix = re.sub(r"^\[[^\]]+\]\s*", "", text).strip()
    inferred = _canonical_ui_category_name(normalize_catalog_category_name(infer_category(name)))
    if _should_repair_catalog_category(current, inferred):
        return True
    if current != "Материнская плата" and re.search(r"^\s*(?:mb|motherboard|мат\s+плат|материнск)\b", text_without_prefix):
        return True
    if current in {"БУМАГА", "АКСЕССУАРЫ"}:
        return True
    if current in {"WEB", "РАЗВЕТВИТЕЛЬ", "НАБОР"}:
        return True
    if current in {"SSD", "Накопители USB", "Монитор"} and (
        "радиатор" in text or "охлажд" in text or "термопаст" in text
        or "web камера" in text or "webcam" in text or "разветвитель usb" in text
        or "usb hub" in text or "dvdrw" in text or "dvd-rw" in text
        or "набор" in text or "ssd" in text or "hdd" in text
        or "жестк" in text or "винчестер" in text
    ):
        return True
    if current in {"Монитор", "Периферия", "Аксессуары"} and (
        "кронштейн" in text
    ):
        return True
    if current == "Кронштейны" and "кронштейн" not in text and (
        "монитор" in text or "ips" in text or "hdmi" in text or "displayport" in text
        or "гц" in text or "hz" in text or re.search(r"\d{3,4}\s*x\s*\d{3,4}", text)
    ):
        return True
    if current == "Кабели и переходники" and (
        "web" in text or "кам" in text or "клавиат" in text or "keyboard" in text
        or "науш" in text or "гарнитур" in text or "wi-fi" in text or "wifi" in text
        or "bluetooth" in text or "сетевой usb" in text
    ):
        return True
    if current == "Периферия" and (
        "монитор" in text or "мышь" in text or "mouse" in text or "клавиат" in text or "keyboard" in text
    ):
        return True
    if current == "Наушники" and (
        "колонки" in text or "акустик" in text or "soundbar" in text or "speaker" in text
    ):
        return True
    if current == "Системный блок" and re.search(r"\bddr[345]\b|оперативн|\bram\b|so[\s\-]?dimm|\bdimm\b", text_without_prefix):
        if not re.search(r"\bкомпьютер\b|системный\s+блок|\bпэвм\b|\btgpc\b|iven\s+(?:by|gaming|office|home|pro|ultra)|\bcore\s+i[3579]\b|\bryzen\b", text_without_prefix):
            return True
    if current in {"27", "24", "32"} and (
        "ips" in text or "hdmi" in text or "displayport" in text or "гц" in text or "hz" in text
    ):
        return True
    if current in {"Видеокарта", "SSD", "Процессор", "Оперативная память"} and (
        "компьютер" in text or "моноблок" in text or "системный блок" in text or "пэвм tgpc" in text
    ):
        return True
    if current == "SSD" and "шасси" in text:
        return True
    return False


def _consolidated_json_df(session_dir, *, apply_visibility=True):
    rows = _correct_consolidated_json_rows(session_dir, apply_visibility=apply_visibility)
    if rows is None:
        return None
    records = []
    indexes = []
    for pos, row in enumerate(rows):
        records.append({
            "OnlinerID": row[0],
            "Название": row[1],
            "Цена": row[2],
            "Поставщик": row[3],
            "Гарантия": row[4],
            "Дней доставки": row[5],
            "РРЦ": row[6],
            "Цена без скидки": row[7],
            "Категория": row[9],
        })
        try:
            indexes.append(int(row[8]))
        except Exception:
            indexes.append(pos)
    return pd.DataFrame(records, index=indexes)


def _precomputed_row_category(row, overrides=None):
    category = str(row.get("Категория", "")).strip()
    return category or row_category(row, overrides)


def api_consolidated():
    session_dir = get_active_session_dir()
    if not session_dir:
        return jsonify({"data": []})
    if not _has_consolidated_session_file(session_dir):
        return jsonify({"data": []})
    try:
        cons_data = _correct_consolidated_json_rows(session_dir, apply_visibility=True)
        if cons_data is not None:
            cons_data = _filter_json_rows_by_export_name_exclusions(cons_data)
            response = jsonify({"data": cons_data})
            response.headers["Cache-Control"] = "no-store, no-cache, must-revalidate, max-age=0"
            response.headers["Pragma"] = "no-cache"
            response.headers["Expires"] = "0"
            return response
    except Exception as exc:
        APP_LOGGER.warning("consolidated JSON fast path failed: %s", exc)
    df = read_consolidated_json_fast_df(session_dir)
    df = ensure_category_column(df)
    df = apply_visibility_filter(df, session_dir)
    df = _filter_df_by_export_name_exclusions(df)
    cons_data = _main_build_consolidated_rows(
        df,
        safe_json_value=_safe_json_value,
        delivery_days_from_row=_delivery_days_from_row,
        row_category=row_category,
    )
    response = jsonify({"data": cons_data})
    response.headers["Cache-Control"] = "no-store, no-cache, must-revalidate, max-age=0"
    response.headers["Pragma"] = "no-cache"
    response.headers["Expires"] = "0"
    return response


def _main_table_badge_counts(rows):
    counts = {
        key: 0
        for key in (
            "cpu", "board", "monitor", "gpu", "ram", "ssd", "psu", "case", "hdd",
            "cooler", "printer", "peripheral", "usb", "network", "ups", "keyboard",
            "mouse", "headphones", "audio", "misc", "ntech-pc", "iven-pc",
            "iven-laptop", "iven-zakaz-laptop", "tradex-laptop",
        )
    }
    core_categories = {
        "Процессор": "cpu",
        "Материнская плата": "board",
        "Монитор": "monitor",
        "Видеокарта": "gpu",
        "Оперативная память": "ram",
        "SSD": "ssd",
        "Блок питания": "psu",
        "Корпус": "case",
        "Жесткий диск": "hdd",
        "Кулер": "cooler",
        "Кулеры": "cooler",
        "Охлаждение": "cooler",
        "Принтер": "printer",
        "Принтеры": "printer",
        "Принтер и МФУ": "printer",
        "МФУ": "printer",
        "Картриджи": "printer",
    }
    for row in rows or []:
        if not isinstance(row, list) or normalize_onliner_id(row[0] if row else ""):
            continue
        name = str(row[1] if len(row) > 1 else "").strip()
        supplier = str(row[3] if len(row) > 3 else "").strip()
        category = normalize_catalog_category_name(str(row[9] if len(row) > 9 else "").strip())
        supplier_token = re.sub(r"[\s-]+", "_", supplier.lower()).strip("_")
        if supplier_token in {"iven", "ивен"} and _is_iven_pc_name(name):
            counts["iven-pc"] += 1
            continue
        if supplier_token in {"iven", "ивен"} and _is_iven_laptop_name(name, category):
            counts["iven-laptop"] += 1
            continue
        if supplier_token in {"iven_zakaz", "ivenzakaz", "ивен_заказ"} and _is_iven_laptop_name(name, category):
            counts["iven-zakaz-laptop"] += 1
            continue
        if supplier_token in {"tradex", "традекс"} and _is_tradex_laptop_name(name, category):
            counts["tradex-laptop"] += 1
            continue
        if supplier_token not in {"n_tech", "ntech", "n_tech"}:
            continue
        if _is_tgpc_pc_name(name):
            counts["ntech-pc"] += 1
            continue
        key = core_categories.get(category)
        if not key:
            for candidate_key, config in (_NTECH_CATEGORY_REVIEW_CONFIG or {}).items():
                if category in set(config.get("categories", set()) or set()):
                    key = candidate_key
                    break
        if key == "cables":
            key = "misc"
        if key not in counts:
            key = "misc"
        counts[key] += 1
    return counts


def api_consolidated_page():
    draw = request.args.get("draw", 0, type=int) or 0
    empty = {"draw": draw, "recordsTotal": 0, "recordsFiltered": 0, "data": [], "meta": {}}
    session_dir = get_active_session_dir()
    if not session_dir or not _has_consolidated_session_file(session_dir):
        return jsonify(empty)
    rows = _correct_consolidated_json_rows(session_dir, apply_visibility=True)
    if rows is None:
        return jsonify(empty)
    rows = _filter_json_rows_by_export_name_exclusions(rows)

    filter_mode = str(request.args.get("filter_mode", "all") or "all").strip().lower()
    export_indexes = set()
    if filter_mode == "export":
        export_df, _ = _prepare_consolidated_for_google_export(session_dir)
        if export_df is not None:
            export_indexes = {str(value) for value in export_df.index.tolist()}
    snapshot_names = set()
    if filter_mode == "snapshot":
        snapshot_mode = str(request.args.get("snapshot_mode", "") or "").strip()
        snapshot_diff = load_session_supplier_diff(session_dir) or {}
        filters = snapshot_diff.get("filters", {}) if isinstance(snapshot_diff, dict) else {}
        key = "new_without_id_names" if snapshot_mode == "new_without_id" else "new_names"
        snapshot_names = set(filters.get(key, []) if isinstance(filters, dict) else [])

    order_specs = []
    for position in range(5):
        column = request.args.get(f"order[{position}][column]")
        if column is None:
            break
        order_specs.append((column, request.args.get(f"order[{position}][dir]", "asc")))
    payload = _build_consolidated_page(
        rows,
        draw=draw,
        start=request.args.get("start", 0, type=int) or 0,
        length=request.args.get("length", 100, type=int) or 100,
        search=request.args.get("search[value]", ""),
        order_specs=order_specs,
        filter_mode=filter_mode,
        no_id_category=request.args.get("no_id_category", ""),
        export_indexes=export_indexes,
        snapshot_names=snapshot_names,
        badge_counts_builder=_main_table_badge_counts,
    )
    response = jsonify(payload)
    response.headers["Cache-Control"] = "no-store, no-cache, must-revalidate, max-age=0"
    return response


def api_stats():
    session_dir = get_active_session_dir()
    if not session_dir:
        return jsonify(_main_empty_stats_payload())
    if not _has_consolidated_session_file(session_dir):
        return jsonify(_main_empty_stats_payload())
    json_rows = _correct_consolidated_json_rows(session_dir, apply_visibility=True)
    if json_rows is not None:
        without_id = 0
        id_counts = {}
        for row in json_rows:
            oid = normalize_onliner_id(row[0] if len(row) > 0 else "")
            if oid:
                id_counts[oid] = int(id_counts.get(oid, 0)) + 1
            else:
                without_id += 1
        quality_payload = _preexport_quality_payload_for_session(session_dir) or {}
        snapshot_diff = load_session_supplier_diff(session_dir)
        new_without_id_count = int((snapshot_diff or {}).get("new_without_id_count", 0) or 0)
        hidden_category_counts = _hidden_category_counts_for_session(session_dir)
        hidden_rows = int(sum(int(item.get("count", 0) or 0) for item in hidden_category_counts))
        return jsonify({
            "without_id": int(without_id),
            "without_id_category_counts": _without_id_category_counts_from_json_rows(json_rows),
            "duplicate_id_rows": int(sum(count for count in id_counts.values() if count > 1)),
            "export_rows": _export_row_count_for_session(session_dir),
            "export_category_counts": _export_category_counts_for_session(session_dir),
            "hidden_rows": hidden_rows,
            "hidden_category_counts": hidden_category_counts,
            "quality_suspicious_price_count": int(quality_payload.get("suspicious_price_count", 0) or 0),
            "new_without_id_count": new_without_id_count,
            "id_pick_badge_count": new_without_id_count if new_without_id_count > 0 else int(without_id),
        })
    df = read_consolidated_df(session_dir)
    payload = _main_build_stats_payload(
        df,
        count_without_onliner_id=_count_rows_without_onliner_id,
        count_duplicate_onliner_id=_count_rows_with_duplicate_onliner_id,
        export_row_count=_export_row_count_for_session(session_dir),
    )
    payload["export_category_counts"] = _export_category_counts_for_session(session_dir)
    payload["without_id_category_counts"] = _without_id_category_counts_from_df(df)
    hidden_category_counts = _hidden_category_counts_for_session(session_dir)
    payload["hidden_rows"] = int(sum(int(item.get("count", 0) or 0) for item in hidden_category_counts))
    payload["hidden_category_counts"] = hidden_category_counts
    snapshot_diff = load_session_supplier_diff(session_dir)
    new_without_id_count = int((snapshot_diff or {}).get("new_without_id_count", 0) or 0)
    payload["new_without_id_count"] = new_without_id_count
    payload["id_pick_badge_count"] = new_without_id_count if new_without_id_count > 0 else int(payload.get("without_id", 0) or 0)
    return jsonify(payload)


def api_export_row_indexes():
    session_dir = get_active_session_dir()
    if not session_dir:
        return jsonify({"indexes": [], "count": 0})
    if not _has_consolidated_session_file(session_dir):
        return jsonify({"indexes": [], "count": 0})
    df, _ = _prepare_consolidated_for_google_export(session_dir)
    if df is None or df.empty:
        return jsonify({"indexes": [], "count": 0})
    indexes = []
    for value in df.index.tolist():
        try:
            indexes.append(int(value))
        except Exception:
            continue
    return jsonify({"indexes": indexes, "count": len(indexes)})


app.register_blueprint(create_main_bp(
    index=index,
    result_page=result_page,
    upload=upload,
    consolidated=api_consolidated,
    consolidated_page=api_consolidated_page,
    stats=api_stats,
    export_row_indexes=api_export_row_indexes,
))


def _get_manual_id_runtime():
    return ManualIdRuntime(
        read_consolidated_json_fast_df=read_consolidated_json_fast_df,
        read_consolidated_df=read_consolidated_df,
        write_consolidated_df=write_consolidated_df,
        write_consolidated_json=write_consolidated_json,
        load_id_cache=load_id_cache,
        save_id_cache=save_id_cache,
        sanitize_id_cache=_sanitize_id_cache,
        load_manual_id_bindings=load_manual_id_bindings,
        save_manual_id_bindings=save_manual_id_bindings,
        load_review_queue=load_review_queue,
        save_review_queue=save_review_queue,
        append_id_change_journal=append_id_change_journal,
        load_id_change_journal=load_id_change_journal,
        save_id_change_journal=save_id_change_journal,
        fetch_onliner_product_info=fetch_onliner_product_info,
        normalize_name_key=_normalize_name_key,
        coerce_bool=_coerce_bool,
        get_id_cache_key_for_name=_get_id_cache_key_for_name,
    )


@_serialized_price_mutation
def _manual_id_confirm_batch_payload(session_dir, payload):
    return _get_manual_id_runtime().confirm_batch(session_dir, payload)


@_serialized_price_mutation
def _manual_id_clear_payload(session_dir, payload):
    return _get_manual_id_runtime().clear(session_dir, payload)


@_serialized_price_mutation
def _manual_id_reject_match_payload(session_dir, payload):
    return _reject_iven_match(
        session_dir,
        payload,
        read_consolidated_df=read_consolidated_df,
        write_consolidated_df=write_consolidated_df,
        write_consolidated_json=write_consolidated_json,
        normalize_name_key=_normalize_name_key,
        load_manual_id_bindings=load_manual_id_bindings,
        save_manual_id_bindings=save_manual_id_bindings,
        blank_id_value=np.nan,
    )


@_serialized_price_mutation
def _autofill_ntech_pc_worker(session_dir, max_items=0):
    return _autofill_run_tgpc_pc(
        session_dir,
        max_items=max_items,
        status=autofill_ntech_pc_status,
        lock=AUTOFILL_NTECH_PC_LOCK,
        read_consolidated_df=read_consolidated_json_fast_df,
        load_app_settings=load_app_settings,
        row_category=row_category,
        is_tgpc_pc_name=_is_tgpc_pc_name,
        db_search_tgpc_pc_candidates=db_search_tgpc_pc_candidates,
        get_id_cache_key_for_name=_get_id_cache_key_for_name,
        normalize_name_key=_normalize_name_key,
        load_id_cache=load_id_cache,
        save_id_cache=save_id_cache,
        load_manual_id_bindings=load_manual_id_bindings,
        save_manual_id_bindings=save_manual_id_bindings,
        append_id_change_journal=append_id_change_journal,
        write_consolidated_df=lambda target, df: write_consolidated_df_background(
            target,
            df,
            label="autofill-ntech-pc",
        ),
        write_consolidated_json=write_consolidated_json,
        target_supplier_names=["N-Tech"],
        pc_label="N-Tech ПЭВМ",
        action_name="autofill_ntech_pc_ids",
        source_name="db_autofill_ntech_pc_ids",
        get_match_identity_for_name=_extract_tgpc_pc_code,
    )


@_serialized_price_mutation
def _autofill_iven_pc_worker(session_dir, max_items=0):
    return _autofill_run_tgpc_pc(
        session_dir,
        max_items=max_items,
        status=autofill_iven_pc_status,
        lock=AUTOFILL_IVEN_PC_LOCK,
        read_consolidated_df=read_consolidated_json_fast_df,
        load_app_settings=load_app_settings,
        row_category=row_category,
        is_tgpc_pc_name=_is_iven_pc_name,
        db_search_tgpc_pc_candidates=db_search_iven_pc_candidates,
        get_id_cache_key_for_name=_get_id_cache_key_for_name,
        normalize_name_key=_normalize_name_key,
        load_id_cache=load_id_cache,
        save_id_cache=save_id_cache,
        load_manual_id_bindings=load_manual_id_bindings,
        save_manual_id_bindings=save_manual_id_bindings,
        append_id_change_journal=append_id_change_journal,
        write_consolidated_df=lambda target, df: write_consolidated_df_background(
            target,
            df,
            label="autofill-iven-pc",
        ),
        write_consolidated_json=write_consolidated_json,
        target_supplier_names=["IVEN"],
        pc_label="IVEN ПЭВМ",
        action_name="autofill_iven_pc_ids",
        source_name="db_autofill_iven_pc_ids",
        get_id_cache_keys_for_name=_id_cache_keys_for_iven_pc_name,
        get_manual_binding_keys_for_name=_manual_binding_keys_for_name,
        get_match_identity_for_name=_extract_iven_pc_code,
        clear_duplicate_ids_for_suppliers=lambda df: _clear_duplicate_onliner_ids_for_suppliers(
            df,
            ["IVEN"],
        ),
    )


def _start_pevm_autofill(status, lock, status_label, worker):
    session_dir = get_active_session_dir()
    payload = request.get_json(silent=True) or {}
    try:
        max_items = int(payload.get("limit", 0) or 0)
    except (TypeError, ValueError):
        max_items = 0
    max_items = max(0, min(max_items, 200))
    return _autofill_start_payload(
        session_dir,
        cons_exists=bool(
            session_dir and _has_consolidated_session_file(session_dir)
        ),
        status=status,
        lock=lock,
        start_worker=lambda: threading.Thread(
            target=worker,
            args=(str(session_dir), max_items),
            daemon=True,
        ).start(),
        status_factory=lambda: _autofill_make_pc_status(status_label),
    )


def api_autofill_ntech_pc_ids():
    return _start_pevm_autofill(
        autofill_ntech_pc_status,
        AUTOFILL_NTECH_PC_LOCK,
        "N-Tech ПЭВМ",
        _autofill_ntech_pc_worker,
    )


def api_autofill_ntech_pc_status():
    return _autofill_status_payload(
        autofill_ntech_pc_status,
        AUTOFILL_NTECH_PC_LOCK,
    )


def api_autofill_iven_pc_ids():
    return _start_pevm_autofill(
        autofill_iven_pc_status,
        AUTOFILL_IVEN_PC_LOCK,
        "IVEN ПЭВМ",
        _autofill_iven_pc_worker,
    )


def api_autofill_iven_pc_status():
    return _autofill_status_payload(
        autofill_iven_pc_status,
        AUTOFILL_IVEN_PC_LOCK,
    )


app.register_blueprint(create_autofill_bp(
    handlers={
        "/api/autofill-ntech-pc-ids": (
            api_autofill_ntech_pc_ids,
            ("POST",),
        ),
        "/api/autofill-ntech-pc-status": (
            api_autofill_ntech_pc_status,
            ("GET",),
        ),
        "/api/autofill-iven-pc-ids": (
            api_autofill_iven_pc_ids,
            ("POST",),
        ),
        "/api/autofill-iven-pc-status": (
            api_autofill_iven_pc_status,
            ("GET",),
        ),
    },
))


def _onliner_db_stats_payload():
    return db_stats()


def _onliner_db_search_payload(query):
    return _onliner_db_search(query)


def _onliner_db_rebuild_payload(session_dir):
    return _onliner_db_rebuild(
        session_dir,
        read_consolidated_df=read_consolidated_df,
        populate_from_dataframe=db_populate_from_df,
    )


def _onliner_db_import_gsheet_payload(payload):
    return _onliner_db_import_gsheet(
        payload,
        normalize_name_key=_normalize_name_key,
        start_thread=lambda target: threading.Thread(target=target, daemon=True).start(),
    )


def _onliner_db_import_csv_payload(file):
    return _onliner_db_import_csv(
        file,
        normalize_name_key=_normalize_name_key,
        start_thread=lambda target: threading.Thread(target=target, daemon=True).start(),
    )


def _onliner_db_import_status_payload():
    return _onliner_db_import_status()


app.register_blueprint(create_onliner_db_bp(
    get_stats=_onliner_db_stats_payload,
    search=_onliner_db_search_payload,
    rebuild=_onliner_db_rebuild_payload,
    import_gsheet=_onliner_db_import_gsheet_payload,
    import_csv=_onliner_db_import_csv_payload,
    get_import_status=_onliner_db_import_status_payload,
    get_active_session_dir=get_active_session_dir,
))


@_serialized_price_mutation
def _manual_id_rollback_last_payload(session_dir):
    return _get_manual_id_runtime().rollback_last(session_dir)


app.register_blueprint(create_manual_id_bp(
    get_active_session_dir=get_active_session_dir,
    confirm_batch=_manual_id_confirm_batch_payload,
    clear=_manual_id_clear_payload,
    reject_match=_manual_id_reject_match_payload,
    rollback_last=_manual_id_rollback_last_payload,
))


def _get_experimental_noid_runtime():
    global EXPERIMENTAL_NOID_RUNTIME
    if EXPERIMENTAL_NOID_RUNTIME is not None:
        return EXPERIMENTAL_NOID_RUNTIME
    with EXPERIMENTAL_NOID_RUNTIME_LOCK:
        if EXPERIMENTAL_NOID_RUNTIME is None:
            EXPERIMENTAL_NOID_RUNTIME = ExperimentalNoIdRuntime(
                db_connection=_onliner_db_connection,
                read_dataframe=read_consolidated_json_fast_df,
                normalize_onliner_id=normalize_onliner_id,
                normalize_name_key=_normalize_name_key,
                find_exact=db_find_exact_id_for_name,
                find_top_candidates=db_find_top_candidates,
                confirm_batch=_manual_id_confirm_batch_payload,
                max_workers=8,
            )
    return EXPERIMENTAL_NOID_RUNTIME


app.register_blueprint(create_experimental_noid_bp(
    get_active_session_dir=get_active_session_dir,
    get_runtime=_get_experimental_noid_runtime,
))


@_serialized_price_mutation
def _clear_invalid_onliner_ids_payload(session_dir, payload):
    return _bulk_clear_invalid_onliner_ids(
        session_dir,
        payload,
        read_consolidated_df=read_consolidated_df,
        write_consolidated_df=write_consolidated_df,
        write_consolidated_json=write_consolidated_json,
        load_id_cache=load_id_cache,
        save_id_cache=save_id_cache,
        load_manual_id_bindings=load_manual_id_bindings,
        save_manual_id_bindings=save_manual_id_bindings,
        build_item_category_key=build_item_category_key,
        normalize_name_key=_normalize_name_key,
        get_id_cache_key_for_name=_get_id_cache_key_for_name,
    )


@_serialized_price_mutation
def _clear_all_nonpc_onliner_ids_payload(session_dir):
    return _bulk_clear_all_nonpc_onliner_ids(
        session_dir,
        read_consolidated_df=read_consolidated_df,
        write_consolidated_df=write_consolidated_df,
        write_consolidated_json=write_consolidated_json,
        append_id_change_journal=append_id_change_journal,
        load_review_queue=load_review_queue,
        save_review_queue=save_review_queue,
        load_manual_id_bindings=load_manual_id_bindings,
        save_manual_id_bindings=save_manual_id_bindings,
        load_id_cache=load_id_cache,
        save_id_cache=save_id_cache,
        is_tgpc_pc_name=_is_tgpc_pc_name,
        normalize_name_key=_normalize_name_key,
        get_id_cache_key_for_name=_get_id_cache_key_for_name,
    )


@_serialized_price_mutation
def _clear_ntech_duplicate_onliner_ids_payload(session_dir):
    return _bulk_clear_ntech_duplicate_onliner_ids(
        session_dir,
        read_consolidated_df=read_consolidated_df,
        write_consolidated_df=write_consolidated_df,
        write_consolidated_json=write_consolidated_json,
        append_id_change_journal=append_id_change_journal,
        load_review_queue=load_review_queue,
        save_review_queue=save_review_queue,
        load_manual_id_bindings=load_manual_id_bindings,
        save_manual_id_bindings=save_manual_id_bindings,
        load_id_cache=load_id_cache,
        save_id_cache=save_id_cache,
        normalize_name_key=_normalize_name_key,
        get_id_cache_key_for_name=_get_id_cache_key_for_name,
    )


app.register_blueprint(create_bulk_id_bp(
    get_active_session_dir=get_active_session_dir,
    clear_invalid=_clear_invalid_onliner_ids_payload,
    clear_all_nonpc=_clear_all_nonpc_onliner_ids_payload,
    clear_ntech_duplicates=_clear_ntech_duplicate_onliner_ids_payload,
))


def _verify_all_ids_one(
    row_idx,
    row,
    settings,
    product_cache,
    manual_bindings,
    product_info_results,
    product_info_events,
    product_info_lock,
):
    def _fetch_info(onliner_id, **kwargs):
        oid = normalize_onliner_id(onliner_id)
        with product_info_lock:
            cached_result = product_info_results.get(oid)
            if cached_result is not None:
                return dict(cached_result)
            event = product_info_events.get(oid)
            owns_request = event is None
            if owns_request:
                event = threading.Event()
                product_info_events[oid] = event
        if owns_request:
            try:
                result = fetch_onliner_product_info(oid, cache=product_cache, **kwargs)
            except Exception:
                result = {"name": "", "url": "", "source": "error"}
            with product_info_lock:
                product_info_results[oid] = dict(result or {})
                product_info_events.pop(oid, None)
                event.set()
            return dict(result or {})
        event.wait(timeout=30)
        with product_info_lock:
            return dict(product_info_results.get(oid) or {})

    def _manual_confirmed(name, onliner_id, supplier_name=""):
        return _manual_store_is_confirmed_id(
            name,
            onliner_id,
            supplier_name=supplier_name,
            load_bindings=lambda: manual_bindings,
            normalize_name_key_func=_normalize_name_key,
        )

    return _id_validation_verify_row(
        row_idx,
        row,
        settings=settings,
        fetch_onliner_product_info=_fetch_info,
        row_category=lambda item: item.get("Категория", ""),
        is_manually_confirmed_id=_manual_confirmed,
        calc_name_match=calc_name_match,
        coerce_bool=_coerce_bool,
        coerce_float=_coerce_float,
    )


def _update_verify_all_ids_status(payload):
    with VERIFY_ALL_IDS_LOCK:
        verify_all_ids_status.update(payload)
        snapshot = dict(verify_all_ids_status)
        snapshot["items"] = list(snapshot.get("items", []) or [])
        snapshot["report_items"] = list(snapshot.get("report_items", []) or [])
    writer = VERIFY_ALL_IDS_STATUS_WRITER
    if callable(writer):
        writer(snapshot)


def _verify_all_ids_worker(session_dir):
    return run_verify_all_worker(session_dir, {
        "status": verify_all_ids_status,
        "status_lock": VERIFY_ALL_IDS_LOCK,
        "status_update": _update_verify_all_ids_status,
        "read_consolidated_df": read_consolidated_json_fast_df,
        "ensure_category_column": ensure_category_column,
        "apply_visibility_filter": apply_visibility_filter,
        "collect_tasks": _id_validation_collect_tasks,
        "is_tgpc_pc_name": _is_tgpc_pc_name,
        "load_app_settings": load_app_settings,
        "load_product_cache": load_onliner_product_cache,
        "load_manual_id_bindings": load_manual_id_bindings,
        "get_max_workers": get_onliner_api_max_workers,
        "verify_one": _verify_all_ids_one,
        "sort_result_items": _id_validation_sort_verify_items,
        "clock": time.time,
    })


def _update_validate_clean_status(payload):
    with VALIDATE_CLEAN_IDS_LOCK:
        validate_clean_ids_status.update(payload)


def _prepare_validate_clean_start(session_dir):
    VALIDATE_CLEAN_CANCEL_EVENT.clear()
    VALIDATE_CLEAN_ANALYSIS_RUNNER.reset_cancel(session_dir)


def _cancel_validate_clean(session_dir):
    VALIDATE_CLEAN_CANCEL_EVENT.set()
    VALIDATE_CLEAN_ANALYSIS_RUNNER.cancel(session_dir)


def _validate_clean_cancel_requested():
    return VALIDATE_CLEAN_CANCEL_EVENT.is_set()


def _raise_if_validate_clean_cancelled():
    if _validate_clean_cancel_requested():
        raise _IdValidationCancelledError("validation cancelled")


def _get_ntech_review_runtime():
    return NTechReviewRuntime(
        get_active_session_dir=get_active_session_dir,
        has_consolidated_session_file=_has_consolidated_session_file,
        consolidated_json_df=_consolidated_json_df,
        read_consolidated_json_fast_df=read_consolidated_json_fast_df,
        ensure_category_column=ensure_category_column,
        precomputed_row_category=_precomputed_row_category,
        row_category=row_category,
        get_handler=_ntech_review_handler,
        load_review_queue=load_review_queue,
        save_review_queue=save_review_queue,
        normalize_catalog_category_name=normalize_catalog_category_name,
        normalize_onliner_id=normalize_onliner_id,
        status=id_review_status,
        status_lock=ID_REVIEW_STATUS_LOCK,
        clock=time.time,
    )


def _ntech_review_queue_start_response(**kwargs):
    result = _get_ntech_review_runtime().start(**kwargs)
    if isinstance(result, tuple) and len(result) == 2:
        payload, status = result
        return jsonify(payload), status
    return jsonify(result)

_NTECH_REVIEW_HANDLERS = None


def _get_ntech_review_handlers():
    global _NTECH_REVIEW_HANDLERS
    if _NTECH_REVIEW_HANDLERS is None:
        _NTECH_REVIEW_HANDLERS = _build_ntech_review_handlers_from_runtime({
            "normalize_name_key": _normalize_name_key,
            "normalize_compact_name": _normalize_compact_name,
            "raw_paren_article_tokens": _raw_paren_article_tokens,
            "is_spec_code": _is_spec_code,
            "db_connection": _db_connection,
            "db_find_top_candidates": db_find_top_candidates,
            "db_find_exact_id_for_name": db_find_exact_id_for_name,
            "infer_category": infer_category,
            "normalize_catalog_category_name": normalize_catalog_category_name,
            "review_cpu_brand_model_key": _review_cpu_brand_model_key,
            "review_looks_like_cpu_name": _review_looks_like_cpu_name,
            "review_find_cpu_candidates": _review_find_cpu_candidates,
            "review_board_brand_model_key": _review_board_brand_model_key,
            "review_find_board_candidates": _review_find_board_candidates,
            "review_monitor_brand_model_key": _review_monitor_brand_model_key,
            "review_find_monitor_candidates": _review_find_monitor_candidates,
            "review_gpu_brand_model_key": _review_gpu_brand_model_key,
            "review_find_gpu_candidates": _review_find_gpu_candidates,
            "review_ram_brand_model_key": _review_ram_brand_model_key,
            "review_find_ram_candidates": _review_find_ram_candidates,
            "review_ssd_brand_model_key": _review_ssd_brand_model_key,
            "review_find_ssd_candidates": _review_find_ssd_candidates,
            "review_psu_brand_model_key": _review_psu_brand_model_key,
            "review_find_psu_candidates": _review_find_psu_candidates,
            "review_case_brand_model_key": _review_case_brand_model_key,
            "review_looks_like_case_name": _review_looks_like_case_name,
            "review_find_case_candidates": _review_find_case_candidates,
            "review_hdd_brand_model_key": _review_hdd_brand_model_key,
            "review_looks_like_hdd_name": _review_looks_like_hdd_name,
            "review_find_hdd_candidates": _review_find_hdd_candidates,
            "review_cooler_brand_model_key": _review_cooler_brand_model_key,
            "review_looks_like_cooler_name": _review_looks_like_cooler_name,
            "review_looks_like_liquid_cpu_cooling_name": _review_looks_like_liquid_cpu_cooling_name,
            "review_find_cooler_candidates": _review_find_cooler_candidates,
            "review_printer_mfp_brand_model_key": _review_printer_mfp_brand_model_key,
            "review_looks_like_printer_or_mfp_name": _review_looks_like_printer_or_mfp_name,
            "review_find_printer_candidates": _review_find_printer_candidates,
            "review_looks_like_peripheral_name": _review_looks_like_peripheral_name,
            "review_find_peripheral_candidates": _review_find_peripheral_candidates,
        })
    return _NTECH_REVIEW_HANDLERS


def _ntech_review_handler(mode):
    return _get_ntech_review_handlers()[mode]




def _validate_clean_ids_worker(session_dir):
    return run_api_validation_worker(session_dir, {
        "mutation_lock": PRICE_DATA_MUTATION_LOCK,
        "read_consolidated_df": read_consolidated_json_fast_df,
        "ensure_category_column": ensure_category_column,
        "load_manual_id_bindings": load_manual_id_bindings,
        "load_review_queue": load_review_queue,
        "collect_tasks": _id_validation_collect_tasks,
        "is_tgpc_pc_name": _is_tgpc_pc_name,
        "load_app_settings": load_app_settings,
        "normalize_onliner_id": normalize_onliner_id,
        "is_manually_confirmed_id": _manual_store_is_confirmed_id,
        "normalize_name_key": _normalize_name_key,
        "build_no_column_state": _id_validation_clean_no_column_state,
        "build_prepare_state": _id_validation_clean_prepare_progress_state,
        "build_no_tasks_state": _id_validation_clean_no_tasks_state,
        "analysis_runner": VALIDATE_CLEAN_ANALYSIS_RUNNER,
        "product_cache_ttl": ONLINER_PRODUCT_CACHE_TTL,
        "get_max_workers": get_onliner_api_max_workers,
        "progress_update": _update_validate_clean_status,
        "raise_if_cancelled": _raise_if_validate_clean_cancelled,
        "cancel_requested": _validate_clean_cancel_requested,
        "apply_api_result": _id_validation_apply_api_result,
        "clear_value": np.nan,
        "populate_review_queue": _id_validation_populate_api_review_queue,
        "save_results": _id_validation_save_clean_results,
        "save_manual_id_bindings": save_manual_id_bindings,
        "save_review_queue": save_review_queue,
        "append_id_change_journal": append_id_change_journal,
        "write_consolidated_df": lambda path, frame: write_consolidated_df_background(
            path,
            frame,
            label="validate-clean-api",
        ),
        "write_consolidated_json": write_consolidated_json,
        "build_finish_state": _id_validation_clean_finish_state,
        "cancelled_error": _IdValidationCancelledError,
        "build_cancelled_state": _id_validation_clean_cancelled_state,
        "build_error_state": _id_validation_clean_error_state,
    })


@_serialized_price_mutation
def _validate_clean_ids_db_worker(session_dir):
    return run_db_validation_worker(session_dir, {
        "status": validate_clean_ids_status,
        "lock": VALIDATE_CLEAN_IDS_LOCK,
        "read_consolidated_df": read_consolidated_df,
        "ensure_category_column": ensure_category_column,
        "load_manual_id_bindings": load_manual_id_bindings,
        "load_review_queue": load_review_queue,
        "collect_tasks": _id_validation_collect_tasks,
        "is_tgpc_pc_name": _is_tgpc_pc_name,
        "build_no_column_state": _id_validation_clean_no_column_state,
        "build_prepare_state": _id_validation_clean_prepare_progress_state,
        "build_no_tasks_state": _id_validation_clean_no_tasks_state,
        "run_db_tasks": _id_validation_run_db_tasks,
        "is_manually_confirmed_id": is_manually_confirmed_id,
        "db_get_product_by_id": db_get_product_by_id,
        "db_find_exact_id_for_name": db_find_exact_id_for_name,
        "calc_name_match": calc_name_match,
        "normalize_name_key": _normalize_name_key,
        "progress_update": _update_validate_clean_status,
        "log": lambda message: APP_LOGGER.info("%s", message),
        "clear_value": np.nan,
        "cancel_requested": _validate_clean_cancel_requested,
        "raise_if_cancelled": _raise_if_validate_clean_cancelled,
        "populate_review_queue": _id_validation_populate_db_review_queue,
        "db_find_top_candidates": db_find_top_candidates,
        "save_results": _id_validation_save_clean_results,
        "save_manual_id_bindings": save_manual_id_bindings,
        "save_review_queue": save_review_queue,
        "append_id_change_journal": append_id_change_journal,
        "write_consolidated_df": write_consolidated_df,
        "write_consolidated_json": write_consolidated_json,
        "build_finish_state": _id_validation_clean_finish_state,
        "cancelled_error": _IdValidationCancelledError,
        "build_cancelled_state": _id_validation_clean_cancelled_state,
        "build_error_state": _id_validation_clean_error_state,
    })

def _manual_id_specialized_candidates(local_name, category="", top_n=12):
    name = str(local_name or "").strip()
    if not name:
        return []
    limit = max(1, int(top_n or 12))
    normalized_category = normalize_catalog_category_name(str(category or infer_category(name) or "").strip())
    prefix_category = _normalized_category_name(name)
    explicit_prefixes = {
        "Веб-камеры", "МФУ", "Клавиатура", "Мышь", "Наушники", "Микрофоны",
        "Акустика", "Накопители USB",
    }
    if prefix_category in explicit_prefixes:
        normalized_category = prefix_category

    if _is_iven_pc_name(name):
        return db_search_iven_pc_candidates(name, limit=limit)
    if _is_tgpc_pc_name(name):
        return db_search_tgpc_pc_candidates(name, limit=limit)
    if _is_iven_laptop_name(name, normalized_category):
        return _supplier_laptop_review_candidates(
            name,
            top_n=limit,
            candidate_filter=_is_iven_laptop_candidate,
            source_label="manual_laptop_db",
        )

    mode_by_category = {
        "Процессор": "cpu",
        "Материнская плата": "board",
        "Монитор": "monitor",
        "Видеокарта": "gpu",
        "Оперативная память": "ram",
        "SSD": "ssd",
        "Блок питания": "psu",
        "Корпус": "case",
        "Жесткий диск": "hdd",
        "Кулер": "cooler",
        "Кулеры": "cooler",
        "Охлаждение": "cooler",
        "Принтер": "printer",
        "Принтеры": "printer",
        "Принтер и МФУ": "printer",
        "МФУ": "printer",
        "Картриджи": "printer",
        "Клавиатура": "peripheral",
        "Мышь": "peripheral",
        "Наушники": "peripheral",
        "Акустика": "peripheral",
    }
    mode = mode_by_category.get(normalized_category)
    if not mode:
        return []
    handler = _ntech_review_handler(mode)
    row = {"Поставщик": "manual"}
    if not handler.is_target(row, name, normalized_category):
        return []
    result = handler.build_row_result(0, row, name, normalized_category, int(time.time())) or {}
    queue_candidates = (result.get("queue_item") or {}).get("candidates") or []
    report_candidates = (result.get("report_item") or {}).get("candidates") or []
    candidates = list(queue_candidates or report_candidates)
    if mode == "peripheral" and normalized_category in {"Клавиатура", "Мышь", "Наушники", "Акустика"}:
        candidates = [
            candidate
            for candidate in candidates
            if normalize_catalog_category_name(
                infer_category(str((candidate or {}).get("name", "") or ""))
            ) == normalized_category
        ]
    return candidates[:limit]


def _id_replace_candidates_payload(payload):
    return _id_reporting_replace_candidates_payload(
        payload,
        settings=load_app_settings(),
        db_get_product_by_id=db_get_product_by_id,
        db_find_top_candidates=db_find_top_candidates,
        db_find_exact_id_for_name=db_find_exact_id_for_name,
        specialized_candidates=_manual_id_specialized_candidates,
        score_candidate=calc_name_match,
        candidate_allowed=_strict_candidate_allowed,
        category_path_hints=_category_path_hints,
        coerce_bool=_coerce_bool,
    )


def _build_duplicate_onliner_id_issues(df):
    return _export_build_duplicate_onliner_id_issues(
        df,
        load_product_cache=load_onliner_product_cache,
        is_manually_confirmed_id=_manual_confirmed_checker(),
        row_category=_precomputed_row_category,
    )


def _manual_confirmed_checker():
    manual_bindings = load_manual_id_bindings()

    def _checker(name, onliner_id, supplier_name=""):
        oid = normalize_onliner_id(onliner_id)
        if not oid:
            return False
        manual = _lookup_manual_binding_for_name(manual_bindings, name, supplier_name)
        if not isinstance(manual, dict) or bool(manual.get("blocked", False)):
            return False
        return normalize_onliner_id(manual.get("id", "")) == oid

    return _checker


def apply_export_duplicate_id_filter(df, supplier_names=None):
    return _export_filter_duplicate_ids(
        df,
        supplier_names=supplier_names,
        load_product_cache=load_onliner_product_cache,
        is_manually_confirmed_id=_manual_confirmed_checker(),
        row_category=_precomputed_row_category,
    )


def apply_export_keep_lowest_price_per_onliner_id(df):
    return _export_filter_keep_lowest_price(df)


def _is_pc_export_row(row):
    return _export_is_pc_export_row(row, is_tgpc_pc_name=_is_tgpc_pc_name, row_category=_precomputed_row_category)


def apply_export_only_pc_filter(df, supplier_names=None):
    return _export_filter_only_pc(
        df,
        supplier_names=supplier_names,
        is_tgpc_pc_name=_is_tgpc_pc_name,
        row_category=_precomputed_row_category,
    )


def _check_duplicate_onliner_ids_payload():
    session_dir = get_active_session_dir()
    if not session_dir:
        return {"status": "error", "message": "Нет активной сессии"}, 400
    if not _has_consolidated_session_file(session_dir):
        return {"status": "error", "message": "Сводный прайс не найден"}, 400

    df = read_consolidated_json_fast_df(session_dir)
    df = ensure_category_column(df)
    df = apply_visibility_filter(df, session_dir)
    return _id_reporting_duplicate_onliner_ids_payload(df, _build_duplicate_onliner_id_issues)


def _start_isolated_verify_all_ids():
    return VERIFY_ALL_IDS_JOB.start(
        get_active_session_dir(),
        _id_validation_build_verify_start_state(),
    )


def _isolated_verify_all_ids_status():
    with VERIFY_ALL_IDS_LOCK:
        fallback = dict(verify_all_ids_status)
        fallback["items"] = list(fallback.get("items", []) or [])
        fallback["report_items"] = list(fallback.get("report_items", []) or [])
    return VERIFY_ALL_IDS_JOB.status(get_active_session_dir(), fallback=fallback)


def _get_id_validation_runtime():
    return IdValidationRuntime(
        get_active_session_dir=get_active_session_dir,
        verify_status=verify_all_ids_status,
        verify_lock=VERIFY_ALL_IDS_LOCK,
        validate_status=validate_clean_ids_status,
        validate_lock=VALIDATE_CLEAN_IDS_LOCK,
        verify_worker=_verify_all_ids_worker,
        validate_api_worker=_validate_clean_ids_worker,
        validate_db_worker=_validate_clean_ids_db_worker,
        thread_factory=threading.Thread,
        isolated_verify_start=_start_isolated_verify_all_ids,
        isolated_verify_status=_isolated_verify_all_ids_status,
        before_validate_start=_prepare_validate_clean_start,
        cancel_validate=_cancel_validate_clean,
    )


def _verify_all_ids_start_payload():
    return _get_id_validation_runtime().verify_all_start()


def _verify_all_ids_status_payload():
    return _get_id_validation_runtime().verify_all_status()


def _validate_clean_ids_start_payload():
    return _get_id_validation_runtime().validate_clean_start()


def _validate_clean_ids_db_start_payload():
    return _get_id_validation_runtime().validate_clean_db_start()


def _validate_clean_ids_status_payload():
    return _get_id_validation_runtime().validate_clean_status()


def _validate_clean_ids_cancel_payload():
    return _get_id_validation_runtime().validate_clean_cancel()


app.register_blueprint(create_id_validation_bp(
    verify_all_start=_verify_all_ids_start_payload,
    get_verify_all_status=_verify_all_ids_status_payload,
    validate_clean_start=_validate_clean_ids_start_payload,
    validate_clean_db_start=_validate_clean_ids_db_start_payload,
    validate_clean_cancel=_validate_clean_ids_cancel_payload,
    get_validate_clean_status=_validate_clean_ids_status_payload,
))


def _start_core_ntech_review(mode):
    return _ntech_review_queue_start_response(
        **_ntech_core_review_start_kwargs(mode)
    )


def api_cpu_review_queue_start():
    return _start_core_ntech_review("cpu")


def api_motherboard_review_queue_start():
    return _start_core_ntech_review("board")


def api_monitor_review_queue_start():
    return _start_core_ntech_review("monitor")


def api_gpu_review_queue_start():
    return _start_core_ntech_review("gpu")


def api_ram_review_queue_start():
    return _start_core_ntech_review("ram")


def api_ssd_review_queue_start():
    return _start_core_ntech_review("ssd")


def api_psu_review_queue_start():
    return _start_core_ntech_review("psu")


def api_case_review_queue_start():
    return _start_core_ntech_review("case")


def api_hdd_review_queue_start():
    return _start_core_ntech_review("hdd")


def api_cooler_review_queue_start():
    return _start_core_ntech_review("cooler")


def api_printer_review_queue_start():
    return _start_core_ntech_review("printer")


def api_peripheral_review_queue_start():
    return _start_core_ntech_review("peripheral")


def _generic_ntech_review_candidates(product_name, top_n=5):
    return _ntech_find_review_candidates(
        product_name,
        top_n=top_n,
        db_find_exact_id_for_name=db_find_exact_id_for_name,
        db_find_top_candidates=db_find_top_candidates,
        normalize_onliner_id=normalize_onliner_id,
    )


def _supplier_laptop_review_candidates(
    product_name,
    top_n=5,
    *,
    candidate_filter=None,
    source_label="laptop_db",
):
    return _ntech_find_review_candidates(
        product_name,
        top_n=top_n,
        db_find_exact_id_for_name=db_find_exact_id_for_name,
        db_find_top_candidates=db_find_top_candidates,
        normalize_onliner_id=normalize_onliner_id,
        candidate_filter=candidate_filter or _is_iven_laptop_candidate,
        source_label=source_label,
    )


def _iven_laptop_review_candidates(product_name, top_n=5):
    return _supplier_laptop_review_candidates(
        product_name,
        top_n=top_n,
        candidate_filter=_is_iven_laptop_candidate,
        source_label="iven_laptop_db",
    )


def _iven_zakaz_laptop_review_candidates(product_name, top_n=5):
    return _supplier_laptop_review_candidates(
        product_name,
        top_n=top_n,
        candidate_filter=_is_iven_laptop_candidate,
        source_label="iven_zakaz_laptop_db",
    )


def _tradex_laptop_review_candidates(product_name, top_n=5):
    return _supplier_laptop_review_candidates(
        product_name,
        top_n=top_n,
        candidate_filter=_is_tradex_laptop_candidate,
        source_label="tradex_laptop_db",
    )


def _build_supplier_laptop_review_handler(
    *,
    supplier_label,
    is_laptop_name,
    candidates_func,
    reason,
    reason_label,
):
    return _ntech_build_laptop_handler(
        supplier_label=supplier_label,
        is_laptop_name=is_laptop_name,
        candidates_func=candidates_func,
        reason=reason,
        reason_label=reason_label,
        normalize_name_key=_normalize_name_key,
        supplier_scoped_review_queue_key=_supplier_scoped_review_queue_key,
    )


def _build_iven_laptop_review_handler():
    return _build_supplier_laptop_review_handler(
        supplier_label="IVEN",
        is_laptop_name=_is_iven_laptop_name,
        candidates_func=_iven_laptop_review_candidates,
        reason="iven_laptop_manual",
        reason_label="Ноутбук IVEN: кандидаты найдены, требуется ручное подтверждение.",
    )


def _build_iven_zakaz_laptop_review_handler():
    return _build_supplier_laptop_review_handler(
        supplier_label="IVEN_zakaz",
        is_laptop_name=_is_iven_laptop_name,
        candidates_func=_iven_zakaz_laptop_review_candidates,
        reason="iven_zakaz_laptop_manual",
        reason_label="Ноутбук IVEN_zakaz: кандидаты найдены, требуется ручное подтверждение.",
    )


def _build_tradex_laptop_review_handler():
    return _build_supplier_laptop_review_handler(
        supplier_label="Tradex",
        is_laptop_name=_is_tradex_laptop_name,
        candidates_func=_tradex_laptop_review_candidates,
        reason="tradex_laptop_manual",
        reason_label="Ноутбук Tradex: кандидаты найдены, требуется ручное подтверждение.",
    )


def _build_generic_ntech_category_review_handler(config):
    return _ntech_build_generic_category_handler(
        config,
        normalize_catalog_category_name=normalize_catalog_category_name,
        normalize_name_key=_normalize_name_key,
        supplier_scoped_review_queue_key=_supplier_scoped_review_queue_key,
        candidates_func=_generic_ntech_review_candidates,
    )


def api_ntech_category_review_queue_start():
    payload = request.get_json(silent=True) or {}
    key = str(payload.get("key", "") or "").strip().lower()
    config = _NTECH_CATEGORY_REVIEW_CONFIG.get(key)
    if not config:
        return jsonify(
            {"status": "error", "message": "РќРµРёР·РІРµСЃС‚РЅР°СЏ РєР°С‚РµРіРѕСЂРёСЏ N-Tech"}
        ), 400
    is_target_row, build_row_result = (
        _build_generic_ntech_category_review_handler(config)
    )
    return _ntech_review_queue_start_response(
        **_ntech_generic_review_start_kwargs(
            key,
            config,
            is_target_row,
            build_row_result,
        )
    )


def _start_supplier_laptop_review(
    supplier,
    report_mode,
    handler_builder,
):
    is_target_row, build_row_result = handler_builder()
    return _ntech_review_queue_start_response(
        **_ntech_laptop_review_start_kwargs(
            supplier,
            report_mode,
            is_target_row,
            build_row_result,
        )
    )


def api_iven_laptop_review_queue_start():
    return _start_supplier_laptop_review(
        "IVEN",
        "iven_laptop",
        _build_iven_laptop_review_handler,
    )


def api_iven_zakaz_laptop_review_queue_start():
    return _start_supplier_laptop_review(
        "IVEN_zakaz",
        "iven_zakaz_laptop",
        _build_iven_zakaz_laptop_review_handler,
    )


def api_tradex_laptop_review_queue_start():
    return _start_supplier_laptop_review(
        "Tradex",
        "tradex_laptop",
        _build_tradex_laptop_review_handler,
    )


def _review_queue_unique_supplier_names(entry):
    return _review_queue_service_supplier_names(entry)


def _manual_binding_id_conflict(manual_bindings, name_key, onliner_id, supplier_names=None):
    return _review_queue_manual_binding_id_conflict(
        manual_bindings,
        name_key,
        onliner_id,
        supplier_names,
        normalize_onliner_id=normalize_onliner_id,
        manual_binding_scoped_key=_manual_binding_scoped_key,
    )


def _df_onliner_id_conflict_for_supplier(df, name_key, onliner_id, supplier_names):
    return _review_queue_dataframe_id_conflict(
        df,
        name_key,
        onliner_id,
        supplier_names,
        normalize_name_key=_normalize_name_key,
        normalize_onliner_id=normalize_onliner_id,
    )


def _supplier_scoped_review_queue_key(name_key, supplier):
    return _review_queue_service_supplier_scoped_key(name_key, supplier)


def _migrate_review_queue_supplier_scope(queue):
    return _review_queue_migrate_supplier_scope(queue)


def _review_queue_match_name_key(queue_key, entry):
    return _review_queue_service_match_name_key(queue_key, entry)


def _get_review_queue_runtime():
    return ReviewQueueRuntime(
        get_active_session_dir=get_active_session_dir,
        load_review_queue=load_review_queue,
        save_review_queue=save_review_queue,
        read_consolidated_json_fast_df=read_consolidated_json_fast_df,
        write_consolidated_json=write_consolidated_json,
        write_consolidated_df_background=write_consolidated_df_background,
        load_manual_id_bindings=load_manual_id_bindings,
        save_manual_id_bindings=save_manual_id_bindings,
        append_id_change_journal=append_id_change_journal,
        normalize_name_key=_normalize_name_key,
        normalize_onliner_id=normalize_onliner_id,
        manual_binding_scoped_key=_manual_binding_scoped_key,
        clock=time.time,
    )


def _review_queue_response(result):
    if isinstance(result, tuple) and len(result) == 2:
        payload, status = result
        return jsonify(payload), status
    return jsonify(result)


def api_review_queue():
    return _review_queue_response(_get_review_queue_runtime().list())


@_serialized_price_mutation
def api_review_queue_pick():
    payload = request.get_json(silent=True) or {}
    return _review_queue_response(_get_review_queue_runtime().pick(payload))


def api_review_queue_clear():
    return _review_queue_response(_get_review_queue_runtime().clear())


app.register_blueprint(create_review_queue_bp(
    start_handlers={
        "/api/cpu-review-queue-start": api_cpu_review_queue_start,
        "/api/motherboard-review-queue-start": api_motherboard_review_queue_start,
        "/api/monitor-review-queue-start": api_monitor_review_queue_start,
        "/api/gpu-review-queue-start": api_gpu_review_queue_start,
        "/api/ram-review-queue-start": api_ram_review_queue_start,
        "/api/ssd-review-queue-start": api_ssd_review_queue_start,
        "/api/psu-review-queue-start": api_psu_review_queue_start,
        "/api/case-review-queue-start": api_case_review_queue_start,
        "/api/hdd-review-queue-start": api_hdd_review_queue_start,
        "/api/cooler-review-queue-start": api_cooler_review_queue_start,
        "/api/printer-review-queue-start": api_printer_review_queue_start,
        "/api/peripheral-review-queue-start": api_peripheral_review_queue_start,
        "/api/ntech-category-review-queue-start": api_ntech_category_review_queue_start,
        "/api/iven-laptop-review-queue-start": api_iven_laptop_review_queue_start,
        "/api/iven-zakaz-laptop-review-queue-start": api_iven_zakaz_laptop_review_queue_start,
        "/api/tradex-laptop-review-queue-start": api_tradex_laptop_review_queue_start,
    },
    list_queue=api_review_queue,
    pick=api_review_queue_pick,
    clear=api_review_queue_clear,
))


def _categories_payload():
    session_dir = get_active_session_dir()
    if not session_dir:
        return {"categories": []}
    if not _has_consolidated_session_file(session_dir):
        return {"categories": []}

    df = _consolidated_json_df(session_dir, apply_visibility=True)
    if df is None:
        df = read_consolidated_df(session_dir)
        df = ensure_category_column(df)
        df = apply_visibility_filter(df, session_dir)
    return _category_reference_categories_payload(
        df,
        normalize_category=_canonical_ui_category_name,
        normalize_onliner_id=normalize_onliner_id,
        category_sort_key=_category_sort_key,
    )


def _category_catalog_payload():
    session_dir = get_active_session_dir()
    overrides = load_category_overrides()
    df = None
    if session_dir and _has_consolidated_session_file(session_dir):
        df = _consolidated_json_df(session_dir, apply_visibility=False)
        if df is None:
            df = read_consolidated_json_fast_df(session_dir)
            df = ensure_category_column(df, overrides)
    payload = _category_reference_catalog_payload(
        priority_categories=CATEGORY_PRIORITY,
        overrides=overrides,
        markups=load_category_markups(),
        df=df,
        row_category=row_category,
        normalize_category=_canonical_ui_category_name,
        is_sorting_review_category=_is_sorting_review_category,
        category_sort_key=_category_sort_key,
    )
    categories = set(payload.get("categories", []) or [])
    categories.update(
        _canonical_visible_onliner_category_name(name)
        for name in db_get_distinct_categories()
    )
    categories = {name for name in categories if name and not _is_sorting_review_category(name)}
    return {"categories": sorted(categories, key=_category_sort_key)}


def _supplier_visibility_known_categories():
    categories = {_canonical_ui_category_name(name) for name in CATEGORY_PRIORITY}
    categories.update(_canonical_visible_onliner_category_name(name) for name in db_get_distinct_categories())
    categories.update(_canonical_ui_category_name(name) for name in load_category_markups().keys())
    return {name for name in categories if name and not _is_sorting_review_category(name)}


def _looks_like_raw_supplier_category(category):
    text = str(category or "").strip()
    if text in {"SSD", "HDD", "ИБП", "БП", "NAS"}:
        return False
    if len(text) < 2:
        return False
    if text != text.upper():
        return False
    return bool(re.fullmatch(r"[A-ZА-ЯЁ0-9\s\-\+/_\.]+", text)) and len(text.split()) <= 2


def _suppliers_payload():
    session_dir = get_active_session_dir()
    if not session_dir:
        return {"suppliers": []}
    if not _has_consolidated_session_file(session_dir):
        return {"suppliers": []}
    df = _consolidated_json_df(session_dir, apply_visibility=False)
    if df is None:
        df = read_consolidated_json_fast_df(session_dir)
    return _category_reference_suppliers_payload(df)


def _supplier_categories_payload(supplier):
    session_dir = get_active_session_dir()
    if not session_dir:
        return {"categories": []}

    if not _has_consolidated_session_file(session_dir):
        return {"categories": []}

    df = _consolidated_json_df(session_dir, apply_visibility=False)
    if df is None:
        df = read_consolidated_json_fast_df(session_dir)
    overrides = load_category_overrides()
    if "Категория" not in df.columns:
        df = ensure_category_column(df, overrides)
    visibility_map = load_visibility_map(session_dir)

    def normalize_supplier_visibility_category(category):
        return _canonical_visible_onliner_category_name(category)

    return _category_reference_supplier_categories_payload(
        df,
        supplier=GLOBAL_VISIBILITY_KEY,
        visibility_map=visibility_map,
        canonical_supplier_name=_canonical_supplier_name,
        normalize_category=normalize_supplier_visibility_category,
        is_sorting_review_category=_is_sorting_review_category,
        category_sort_key=_category_sort_key,
        include_category=lambda name: not _looks_like_raw_supplier_category(name),
    )


@app.route("/api/onliner-category-preview")
def api_onliner_category_preview():
    session_dir = get_active_session_dir()
    if not session_dir or not _has_consolidated_session_file(session_dir):
        return jsonify({"status": "error", "message": "Нет активной сессии"}), 400

    df = _consolidated_json_df(session_dir, apply_visibility=False)
    if df is None:
        df = read_consolidated_json_fast_df(session_dir)
    if df is None or df.empty:
        return jsonify({"status": "ok", "summary": {}, "categories": [], "transitions": []})

    unique_ids = _onliner_category_preview_collect_ids(df, normalize_onliner_id=normalize_onliner_id)
    catalog_categories = db_get_categories_by_ids(unique_ids)
    return jsonify(_onliner_category_preview_payload(
        df,
        catalog_categories=catalog_categories,
        markups=load_category_markups(),
        normalize_onliner_id=normalize_onliner_id,
        normalize_catalog_category_name=normalize_catalog_category_name,
    ))


app.register_blueprint(create_category_reference_bp(
    get_categories=_categories_payload,
    get_category_catalog=_category_catalog_payload,
    get_suppliers=_suppliers_payload,
    get_supplier_categories=_supplier_categories_payload,
))


def _get_category_management_runtime():
    return CategoryManagementRuntime(
        get_active_session_dir=get_active_session_dir,
        canonical_supplier_name=_canonical_supplier_name,
        load_visibility_map=load_visibility_map,
        save_visibility_map=save_visibility_map,
        update_category_visibility=_category_update_visibility,
        category_sort_key=_category_sort_key,
        visibility_lock=CATEGORY_VISIBILITY_LOCK,
        has_consolidated_session_file=_has_consolidated_session_file,
        consolidated_json_df=_consolidated_json_df,
        read_consolidated_json_fast_df=read_consolidated_json_fast_df,
        read_consolidated_df=read_consolidated_df,
        ensure_category_column=ensure_category_column,
        apply_visibility_filter=apply_visibility_filter,
        parse_markup_request=_category_parse_markup_request,
        apply_markup_to_df=_category_apply_markup_to_df,
        markup_preview_payload=_category_markup_preview_payload,
        row_category=_precomputed_row_category,
        normalize_onliner_id=normalize_onliner_id,
        get_onliner_market_stats_bulk=get_onliner_market_stats_bulk,
        write_consolidated_json=write_consolidated_json,
        write_consolidated_df_background=write_consolidated_df_background,
        load_category_markups=load_category_markups,
        save_category_markups=save_category_markups,
        update_category_markups=_category_update_markups,
        load_category_overrides=load_category_overrides,
        save_category_overrides=save_category_overrides,
        load_manual_category_overrides=load_manual_category_overrides,
        save_manual_category_overrides=save_manual_category_overrides,
        category_override_items_payload=_category_override_items_payload,
        apply_category_override_to_df=_category_apply_override_to_df,
        canonical_ui_category_name=_canonical_ui_category_name,
        build_item_category_key=build_item_category_key,
        build_item_category_keys=build_item_category_keys,
        infer_category=infer_category,
        write_consolidated_df=write_consolidated_df,
        override_lock=CATEGORY_OVERRIDE_LOCK,
        category_preview_items_payload=_category_preview_items_payload,
        load_market_cache=load_onliner_market_cache,
        get_market_stats_from_cache_only=get_onliner_market_stats_from_cache_only,
    )


def api_category_visibility():
    result, status = _get_category_management_runtime().visibility(
        request.get_json(silent=True) or {}
    )
    return jsonify(result), status


@_serialized_price_mutation
def api_apply_markup():
    return jsonify(
        _get_category_management_runtime().apply_markup(
            request.get_json(silent=True) or {}
        )
    )


def api_markup_preview():
    return jsonify(
        _get_category_management_runtime().markup_preview(
            request.get_json(silent=True) or {}
        )
    )


def api_category_override_items():
    return jsonify(
        _get_category_management_runtime().category_override_items(
        query=request.args.get("q", ""),
        limit=request.args.get("limit", 40),
        )
    )


@_serialized_price_mutation
def api_category_override_set():
    return jsonify(
        _get_category_management_runtime().category_override_set(
            request.get_json(silent=True) or {}
        )
    )


def api_category_preview_items():
    return jsonify(
        _get_category_management_runtime().category_preview_items(
            request.get_json(silent=True) or {}
        )
    )


app.register_blueprint(create_category_management_bp(
    handlers={
        "/api/category-visibility": (api_category_visibility, ("POST",)),
        "/api/apply-markup": (api_apply_markup, ("POST",)),
        "/api/markup-preview": (api_markup_preview, ("POST",)),
        "/api/category-override-items": (api_category_override_items, ("GET",)),
        "/api/category-override-set": (api_category_override_set, ("POST",)),
        "/api/category-preview-items": (api_category_preview_items, ("POST",)),
    }
))


def _market_refresh_worker(session_dir, categories):
    return _get_onliner_market_runtime().market_refresh_worker(session_dir, categories)


def _collect_known_onliner_ids(max_ids=AUTO_REFRESH_MAX_IDS, session_dir=None):
    return _get_onliner_market_runtime().collect_known_onliner_ids(max_ids=max_ids, session_dir=session_dir)


def _market_id_hints_from_session(session_dir):
    """Первая строка прайса по каждому OnlinerID → подсказки для B2B-поиска раздела/производителя."""
    return _get_onliner_market_runtime().market_id_hints_from_session(session_dir)


def _auto_market_refresh_loop():
    return _get_onliner_market_runtime().auto_market_refresh_loop()


def api_onliner_b2b_test():
    return _onliner_diag_b2b_test(
        get_token=onliner_b2b_get_token,
        b2b_request=onliner_b2b_request,
    )


def api_onliner_b2b_probe():
    return _onliner_diag_b2b_probe(
        get_token=onliner_b2b_get_token,
        b2b_request=onliner_b2b_request,
    )


def _get_source_runtime():
    return SourceRuntime(
        session_obj=session,
        load_settings=load_app_settings,
        fetch_worker=_fetch_api_source_worker,
        thread_factory=lambda target: threading.Thread(target=target, daemon=True).start(),
        get_history=get_api_fetch_history,
        process_supplier_files=_process_supplier_files,
        finalize_processed_session=_finalize_processed_session,
        append_history=append_api_fetch_history,
        redirect_for_session=lambda sid: url_for("main_api.result_page", sid=sid),
        enqueue_fetch=(
            _enqueue_api_source_fetch
            if DURABLE_JOB_QUEUE is not None
            else None
        ),
    )


def _enqueue_api_source_fetch(source_key, client_key):
    if DURABLE_JOB_QUEUE is None:
        raise RuntimeError("durable job queue is not configured")
    runtime = _api_sources_get_runtime(source_key, client_key)
    job_id = str(runtime.get("job_id", "") or "").strip()
    DURABLE_JOB_QUEUE.enqueue(
        "api_source_fetch",
        {
            "source_key": str(source_key),
            "client_key": str(client_key),
        },
        dedupe_key=f"{client_key}:{source_key}",
        max_attempts=2,
        job_id=job_id or None,
    )


def api_source_fetch_start():
    return _get_source_runtime().fetch_start(request.get_json(silent=True) or {})


def api_source_fetch_status():
    return _get_source_runtime().fetch_status(request.args.get("source", ""))


def api_source_process():
    return _get_source_runtime().process(request.get_json(silent=True) or {})


def api_source_process_batch():
    return _get_source_runtime().process_batch(request.get_json(silent=True) or {})


app.register_blueprint(create_source_bp(
    handlers={
        "/api/source-fetch-start": (api_source_fetch_start, ("POST",)),
        "/api/source-fetch-status": (api_source_fetch_status, ("GET",)),
        "/api/source-process": (api_source_process, ("POST",)),
        "/api/source-process-batch": (api_source_process_batch, ("POST",)),
    }
))


def _start_market_refresh(session_dir, categories):
    return _market_refresh_start(session_dir, categories, worker=_market_refresh_worker)


def _get_market_refresh_status_snapshot():
    return _market_refresh_status_snapshot()


app.register_blueprint(create_settings_bp(
    get_active_session_dir=get_active_session_dir,
    get_last_active_session_dir=lambda: LAST_ACTIVE_SESSION_DIR,
    start_market_refresh=_start_market_refresh,
    save_app_settings=save_app_settings,
))


app.register_blueprint(create_market_bp(
    get_active_session_dir=get_active_session_dir,
    start_market_refresh=_start_market_refresh,
    get_market_refresh_status=_get_market_refresh_status_snapshot,
))


def api_onliner_offers(onliner_id):
    return _onliner_diag_offers(
        onliner_id,
        normalize_onliner_id=normalize_onliner_id,
        fetch_product_payload=_fetch_onliner_product_payload,
        api_get=onliner_api_get,
        extract_offer_rows=_extract_offer_rows,
    )


app.register_blueprint(create_onliner_bp(
    b2b_test=api_onliner_b2b_test,
    b2b_probe=api_onliner_b2b_probe,
    get_offers=api_onliner_offers,
))


def _get_category_extra_runtime():
    return CategoryExtraRuntime(
        get_active_session_dir=get_active_session_dir,
        resolve_session_dir=_resolve_session_dir,
        has_consolidated_session_file=_has_consolidated_session_file,
        lock=PRICE_DATA_MUTATION_LOCK,
        load_category_markups=load_category_markups,
        load_category_overrides=load_category_overrides,
        save_category_overrides=save_category_overrides,
        load_manual_category_overrides=load_manual_category_overrides,
        save_manual_category_overrides=save_manual_category_overrides,
        read_consolidated_json_fast_df=read_consolidated_json_fast_df,
        ensure_category_column=ensure_category_column,
        apply_visibility_filter=apply_visibility_filter,
        apply_saved_markups_to_df=apply_saved_markups_to_df,
        write_consolidated_df=write_consolidated_df,
        write_consolidated_json=write_consolidated_json,
        apply_category_override_to_df=_category_apply_override_to_df,
        autosort_preview_payload=_autosort_preview_payload,
        autosort_apply_items=_autosort_apply_items,
        canonical_ui_category_name=_canonical_ui_category_name,
        build_item_category_keys=build_item_category_keys,
        build_item_category_key=build_item_category_key,
        row_category=row_category,
        name_tokens=_name_tokens,
        normalize_onliner_id=normalize_onliner_id,
        category_sort_key=_category_sort_key,
        predict_openai_category=_openai_autosort_predict_category,
        openai_api_key=OPENAI_API_KEY,
        autosort_max_items=OPENAI_AUTOSORT_MAX_ITEMS,
        autosort_max_workers=OPENAI_AUTOSORT_MAX_WORKERS,
    )


def api_category_markups():
    return jsonify(_get_category_extra_runtime().markups())


def api_category_override_bulk():
    result = _get_category_extra_runtime().override_bulk(request.get_json(silent=True) or {})
    if isinstance(result, tuple):
        payload, status = result
        return jsonify(payload), status
    return jsonify(result)


def api_category_autosort_preview():
    return jsonify(_get_category_extra_runtime().autosort_preview(request.get_json(silent=True) or {}))


@_serialized_price_mutation
def api_category_autosort_apply():
    return jsonify(_get_category_extra_runtime().autosort_apply(request.get_json(silent=True) or {}))


def api_resolve_start():
    return jsonify(_resolve_start_payload(
        resolve_status,
        session_dir=get_active_session_dir(),
        has_consolidated_session_file=_has_consolidated_session_file,
        read_consolidated_json_fast_df=read_consolidated_json_fast_df,
        load_url_cache=load_url_cache,
        resolve_onliner_urls=resolve_onliner_urls,
        read_consolidated_df=read_consolidated_df,
        write_consolidated_df=write_consolidated_df,
        write_consolidated_json=write_consolidated_json,
        thread_factory=lambda target: threading.Thread(
            target=lambda: _serialized_price_mutation(target)(),
            daemon=True,
        ).start(),
    ))


def api_resolve_status():
    return jsonify(_resolve_status_snapshot(resolve_status))


app.register_blueprint(create_resolve_bp(
    start=api_resolve_start,
    status=api_resolve_status,
))


def _resolve_service_account_json_path(raw):
    """
    Ищем JSON ключ сервисного аккаунта. Возвращает (Path|None, подсказка_для_ошибки).
    """
    return _export_resolve_service_account_json_path(
        raw,
        base_dir=Path(__file__).resolve().parent,
        cwd=Path.cwd(),
    )


def _export_settings_with_onliner_structure():
    settings = load_app_settings()
    export_cfg = dict((settings or {}).get("export") or {})
    export_cfg["allowed_categories"] = [
        _canonical_ui_category_name(category)
        for category in EXPORT_ONLINER_STRUCTURE_CATEGORIES
        if _canonical_ui_category_name(category)
    ]
    export_cfg["exclude_category_prefixes"] = _merge_export_text_list(
        export_cfg.get("exclude_category_prefixes", []),
        EXPORT_MANDATORY_CATEGORY_PREFIX_EXCLUDES,
    )
    export_cfg["exclude_name_contains"] = _merge_export_text_list(
        export_cfg.get("exclude_name_contains", []),
        EXPORT_MANDATORY_NAME_EXCLUDES,
    )
    settings = dict(settings or {})
    settings["export"] = export_cfg
    return settings


def _merge_export_text_list(*groups):
    items = []
    seen = set()
    for group in groups:
        raw_items = group if isinstance(group, list) else str(group or "").replace("\r", "\n").split("\n")
        for item in raw_items:
            text = str(item or "").strip()
            key = text.casefold()
            if text and key not in seen:
                seen.add(key)
                items.append(text)
    return items


def _export_name_exclude_patterns():
    export_cfg = (_export_settings_with_onliner_structure() or {}).get("export", {})
    patterns = []
    for item in export_cfg.get("exclude_name_contains", []):
        pattern = str(item or "").strip().casefold()
        if pattern and pattern not in patterns:
            patterns.append(pattern)
    return patterns


def _filter_json_rows_by_export_name_exclusions(rows):
    patterns = _export_name_exclude_patterns()
    if not patterns:
        return rows
    filtered = []
    for row in rows or []:
        name = str(row[1] if len(row) > 1 else "").strip().casefold()
        if name and any(pattern in name for pattern in patterns):
            continue
        filtered.append(row)
    return filtered


def _filter_df_by_export_name_exclusions(df):
    patterns = _export_name_exclude_patterns()
    if not patterns or df is None or getattr(df, "empty", True) or "Название" not in df.columns:
        return df
    name_text = df["Название"].fillna("").astype(str).str.casefold()
    mask = ~name_text.apply(lambda value: any(pattern in value for pattern in patterns))
    return df[mask].copy()


def _quality_stats_cache_key(session_dir):
    base_dir = Path(__file__).resolve().parent
    session_path = Path(session_dir)
    paths = [
        session_path / "consolidated.json",
        session_path / "consolidated_price.xlsx",
        base_dir / "manual_id_bindings.json",
        base_dir / "onliner_market_cache.json",
        base_dir / "app_settings.json",
    ]
    mtimes = []
    for path in paths:
        try:
            mtimes.append(int(path.stat().st_mtime_ns))
        except OSError:
            mtimes.append(0)
    return (str(session_path.resolve()), *mtimes, _category_state_signature())


def _preexport_quality_payload_for_session(session_dir):
    key = _quality_stats_cache_key(session_dir)
    cached = QUALITY_STATS_CACHE.get(key)
    if cached is not None:
        return dict(cached)

    df, _ = _prepare_consolidated_for_google_export(session_dir)
    if df is None:
        return None
    df = ensure_category_column(df)
    payload = _export_build_preexport_quality_payload(df)
    if len(QUALITY_STATS_CACHE) > 12:
        QUALITY_STATS_CACHE.clear()
    QUALITY_STATS_CACHE[key] = dict(payload)
    return payload


def _prepare_consolidated_for_export(session_dir):
    """Тот же набор фильтров, что и для /download. Возвращает (DataFrame|None, имя_файла.xlsx)."""
    return _export_prepare_consolidated(
        session_dir,
        _export_settings_with_onliner_structure(),
        read_consolidated_df=read_consolidated_export_df,
        apply_visibility_filter=apply_visibility_filter,
        apply_keep_lowest_price_per_onliner_id=apply_export_keep_lowest_price_per_onliner_id,
        apply_duplicate_id_filter=apply_export_duplicate_id_filter,
        apply_only_pc_filter=apply_export_only_pc_filter,
    )


def _prepare_consolidated_for_google_export(session_dir):
    """Prepare only the fixed Google export columns from the fast JSON snapshot."""
    key = _quality_stats_cache_key(session_dir)
    cached = GOOGLE_EXPORT_DF_CACHE.get(key)
    if cached is not None:
        cached_df, cached_name = cached
        return cached_df.copy(), cached_name

    prepared = _export_prepare_consolidated(
        session_dir,
        _export_settings_with_onliner_structure(),
        read_consolidated_df=read_consolidated_export_df,
        apply_visibility_filter=apply_visibility_filter,
        apply_keep_lowest_price_per_onliner_id=apply_export_keep_lowest_price_per_onliner_id,
        apply_duplicate_id_filter=apply_export_duplicate_id_filter,
        apply_only_pc_filter=apply_export_only_pc_filter,
    )
    prepared_df, prepared_name = prepared
    if prepared_df is not None:
        if len(GOOGLE_EXPORT_DF_CACHE) > 4:
            GOOGLE_EXPORT_DF_CACHE.clear()
        current_key = _quality_stats_cache_key(session_dir)
        GOOGLE_EXPORT_DF_CACHE[current_key] = (prepared_df.copy(), prepared_name)
    return prepared


def _export_row_count_for_session(session_dir):
    """Return the row count that the current export/Google Sheets pipeline will emit."""
    if not session_dir:
        return 0
    try:
        rows = _correct_consolidated_json_rows(session_dir, apply_visibility=True)
        if rows is not None:
            return _export_row_count_from_json_rows(rows, _export_settings_with_onliner_structure())
    except Exception as exc:
        APP_LOGGER.warning("export row count from JSON failed: %s", exc)
    try:
        filtered, _ = _prepare_consolidated_for_google_export(session_dir)
        return int(len(filtered)) if filtered is not None else 0
    except Exception as exc:
        APP_LOGGER.warning("export row count dataframe fallback failed: %s", exc)
    return 0


def _export_category_counts_for_session(session_dir):
    """Return category counts for the rows that the export/Google Sheets pipeline will emit."""
    if not session_dir:
        return []
    try:
        rows = _correct_consolidated_json_rows(session_dir, apply_visibility=True)
        if rows is not None:
            return _export_category_counts_from_json_rows(rows, _export_settings_with_onliner_structure())
    except Exception as exc:
        APP_LOGGER.warning("export category counts from JSON failed: %s", exc)
    try:
        filtered, _ = _prepare_consolidated_for_google_export(session_dir)
        if filtered is None or filtered.empty:
            return []
        counts = {}
        for category in filtered.get("Категория", pd.Series(dtype=str)).fillna("").astype(str):
            name = normalize_catalog_category_name(str(category or "").strip()) or "Без категории"
            counts[name] = int(counts.get(name, 0)) + 1
        return _format_export_category_counts(counts)
    except Exception as exc:
        APP_LOGGER.warning(
            "export category counts dataframe fallback failed: %s", exc
        )
    return []


def _export_row_count_from_json_rows(json_rows, settings):
    """Fast count for the same rows that /download and Google Sheets export will keep."""
    return _export_stats_row_count_from_json_rows(
        json_rows,
        settings,
        normalize_onliner_id=normalize_onliner_id,
        normalize_name_key=_normalize_name_key,
        normalize_supplier_name_list=_export_normalize_supplier_name_list,
        is_pc_export_row=_is_pc_export_row,
    )


def _export_category_counts_from_json_rows(json_rows, settings):
    return _export_stats_category_counts_from_json_rows(
        json_rows,
        settings,
        normalize_onliner_id=normalize_onliner_id,
        normalize_name_key=_normalize_name_key,
        normalize_supplier_name_list=_export_normalize_supplier_name_list,
        is_pc_export_row=_is_pc_export_row,
        category_sort_key=_category_sort_key,
    )


def _without_id_category_counts_from_json_rows(json_rows):
    return _export_stats_without_id_category_counts_from_json_rows(
        json_rows,
        normalize_onliner_id=normalize_onliner_id,
        category_sort_key=_category_sort_key,
    )


def _without_id_category_counts_from_df(df):
    return _export_stats_without_id_category_counts_from_df(
        df,
        normalize_onliner_id=normalize_onliner_id,
        category_sort_key=_category_sort_key,
    )


def _format_without_id_category_counts(counts):
    return _export_stats_format_category_counts(counts, category_sort_key=_category_sort_key)


def _format_export_category_counts(counts):
    return _export_stats_format_category_counts(counts, category_sort_key=_category_sort_key)


def _hidden_category_counts_from_json_rows(json_rows, session_dir):
    visibility_map = load_visibility_map(session_dir) if session_dir else {}
    if not visibility_map:
        return []
    hidden_categories = {
        _canonical_ui_category_name(normalize_catalog_category_name(category))
        for categories in visibility_map.values()
        for category in categories or []
        if _canonical_ui_category_name(normalize_catalog_category_name(category))
    }
    if not hidden_categories:
        return []
    counts = {}
    for row in json_rows or []:
        if not isinstance(row, list):
            continue
        category = str(row[9] if len(row) > 9 else "").strip()
        if not category:
            category = "Без категории"
        normalized_category = _canonical_ui_category_name(normalize_catalog_category_name(category))
        if normalized_category in hidden_categories:
            counts[normalized_category] = int(counts.get(normalized_category, 0)) + 1
    return _format_export_category_counts(counts)


def _hidden_category_counts_for_session(session_dir):
    if not session_dir:
        return []
    try:
        rows = _correct_consolidated_json_rows(session_dir, apply_visibility=False)
        if rows is not None:
            return _hidden_category_counts_from_json_rows(rows, session_dir)
    except Exception as exc:
        APP_LOGGER.warning("hidden category counts failed: %s", exc)
    return []


def _hidden_row_count_for_session(session_dir):
    return int(sum(int(item.get("count", 0) or 0) for item in _hidden_category_counts_for_session(session_dir)))


def download():
    settings = load_app_settings()
    export_cfg = settings.get("export", {})
    base_name = str(export_cfg.get("price_name", "consolidated_price")).strip() or "consolidated_price"
    download_name = f"{base_name}.xlsx"
    session_dir = get_active_session_dir()
    if session_dir:
        filtered, download_name = _prepare_consolidated_for_export(session_dir)
        if filtered is not None:
            visible_path = Path(session_dir) / "consolidated_price_visible.xlsx"
            export_df = _export_dataframe_to_xlsx(filtered)
            export_df.to_excel(visible_path, index=False, float_format="%.2f")
            return send_file(str(visible_path), as_attachment=True, download_name=download_name)
        output_path = Path(session_dir) / "consolidated_price.xlsx"
        if output_path.exists():
            return send_file(str(output_path), as_attachment=True, download_name=download_name)
    return redirect(url_for("main_api.index", error="Файл не найден. Загрузите прайсы заново."))


def _lookup_onliner_db_products(onliner_ids):
    ids = []
    seen = set()
    for value in onliner_ids or []:
        oid = normalize_onliner_id(value)
        if oid and oid not in seen:
            seen.add(oid)
            ids.append(oid)
    if not ids:
        return {}

    products = {}
    try:
        with _onliner_db_connection() as conn:
            for start in range(0, len(ids), 900):
                chunk = ids[start:start + 900]
                placeholders = ",".join("?" for _ in chunk)
                rows = conn.execute(
                    "SELECT onliner_id, name, url, source, updated_at "
                    f"FROM onliner_catalog WHERE onliner_id IN ({placeholders})",
                    chunk,
                ).fetchall()
                for row in rows:
                    oid = normalize_onliner_id(row["onliner_id"])
                    if not oid:
                        continue
                    products[oid] = {
                        "id": oid,
                        "name": str(row["name"] or "").strip(),
                        "url": str(row["url"] or "").strip(),
                        "source": str(row["source"] or "").strip(),
                        "updated_at": int(row["updated_at"] or 0),
                    }
    except Exception as exc:
        APP_LOGGER.warning("ID compare DB lookup failed: %s", exc)
    return products


def _build_id_compare_report_df(session_dir):
    df = _consolidated_json_df(session_dir, apply_visibility=False)
    if df is None:
        df = read_consolidated_df(session_dir)
    df = ensure_category_column(df)
    if df is None or df.empty or "OnlinerID" not in df.columns:
        return pd.DataFrame()

    ids = [normalize_onliner_id(value) for value in df.get("OnlinerID", pd.Series(dtype=str)).tolist()]
    products_by_id = _lookup_onliner_db_products(ids)
    return _id_compare_report_build_df(
        df,
        products_by_id,
        normalize_onliner_id=normalize_onliner_id,
        calc_name_match=calc_name_match,
    )


def _format_id_compare_workbook(path):
    try:
        from openpyxl import load_workbook
        from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
        from openpyxl.utils import get_column_letter
    except Exception:
        return

    wb = load_workbook(path)
    ws = wb.active
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    ws.sheet_view.showGridLines = True
    header_fill = PatternFill("solid", fgColor="063B46")
    even_fill = PatternFill("solid", fgColor="BDE7F6")
    odd_fill = PatternFill("solid", fgColor="FFFFFF")
    ok_fill = PatternFill("solid", fgColor="D8FBE7")
    warn_fill = PatternFill("solid", fgColor="FFF2CC")
    bad_fill = PatternFill("solid", fgColor="FCE4D6")
    border = Border(
        left=Side(style="thin", color="4AB6D8"),
        right=Side(style="thin", color="4AB6D8"),
        top=Side(style="thin", color="4AB6D8"),
        bottom=Side(style="thin", color="4AB6D8"),
    )
    for cell in ws[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = border
    ws.row_dimensions[1].height = 24
    decision_col = 1
    score_col = 2
    check_col = 11
    price_col = 9
    for row_idx in range(2, ws.max_row + 1):
        row_fill = even_fill if row_idx % 2 == 0 else odd_fill
        for cell in ws[row_idx]:
            cell.fill = row_fill
            cell.border = border
            cell.alignment = Alignment(vertical="top", wrap_text=False)
        decision = str(ws.cell(row=row_idx, column=decision_col).value or "")
        decision_fill = ok_fill if "верный" in decision.lower() else (bad_fill if "не найден" in decision.lower() else warn_fill)
        ws.cell(row=row_idx, column=decision_col).fill = decision_fill
        ws.cell(row=row_idx, column=decision_col).font = Font(bold=True, color="0F5132" if "верный" in decision.lower() else "7C2D12")
        for col_idx in (4, 6, 12):
            ws.cell(row=row_idx, column=col_idx).alignment = Alignment(wrap_text=True, vertical="top")
        for col_idx in (decision_col, score_col, 3, 5, 7, 8, price_col, check_col):
            ws.cell(row=row_idx, column=col_idx).alignment = Alignment(horizontal="center", vertical="top", wrap_text=False)
        try:
            ws.cell(row=row_idx, column=score_col).number_format = "0"
            ws.cell(row=row_idx, column=price_col).number_format = "0.00"
        except Exception:
            pass
    widths = {
        1: 22, 2: 12, 3: 14, 4: 62, 5: 14, 6: 56, 7: 20,
        8: 16, 9: 12, 10: 16, 11: 22, 12: 58, 13: 20, 14: 38,
    }
    for col_idx, width in widths.items():
        ws.column_dimensions[get_column_letter(col_idx)].width = width
    ws.auto_filter.ref = f"A1:{get_column_letter(ws.max_column)}{ws.max_row}"
    wb.save(path)


def download_id_compare_report():
    session_dir = get_active_session_dir()
    if not session_dir:
        return redirect(url_for("main_api.index", error="Нет активной сессии"))
    if not _has_consolidated_session_file(session_dir):
        return redirect(url_for("main_api.index", error="Сводный прайс не найден"))

    report_df = _build_id_compare_report_df(session_dir)
    if report_df.empty:
        return redirect(url_for("main_api.index", error="Нет товаров с OnlinerID для сверки"))
    report_path = Path(session_dir) / "id_compare_report.xlsx"
    with pd.ExcelWriter(report_path, engine="openpyxl") as writer:
        report_df.to_excel(writer, sheet_name="Сверка ID", index=False)
    _format_id_compare_workbook(report_path)
    return send_file(str(report_path), as_attachment=True, download_name="id_compare_report.xlsx")


def api_export_google_sheets():
    return _export_google_sheets_payload(
        get_active_session_dir(),
        load_app_settings(),
        prepare_consolidated_for_export=_prepare_consolidated_for_google_export,
        resolve_service_account_json_path_func=_resolve_service_account_json_path,
    )


def download_id_quality_report():
    session_dir = get_active_session_dir()
    if not session_dir:
        return redirect(url_for("main_api.index", error="Нет активной сессии"))
    path = Path(session_dir) / "id_quality_report.csv"
    if not path.exists():
        return redirect(url_for("main_api.index", error="ID quality report не найден"))
    return send_file(str(path), as_attachment=True, download_name="id_quality_report.csv")


def _id_quality_report_payload():
    session_dir = get_active_session_dir()
    if not session_dir:
        return {"status": "no_session"}
    summary_path = Path(session_dir) / "id_quality_report.json"
    report_path = Path(session_dir) / "id_quality_report.csv"
    if not summary_path.exists():
        return {"status": "not_found"}
    summary = load_dict(summary_path)
    if not summary:
        return {"status": "error"}
    summary["status"] = "ok"
    summary["has_csv"] = report_path.exists()
    return summary


app.register_blueprint(create_id_reporting_bp(
    replace_candidates=_id_replace_candidates_payload,
    check_duplicate_ids=_check_duplicate_onliner_ids_payload,
    get_quality_report=_id_quality_report_payload,
))


def api_preexport_quality_check():
    session_dir = get_active_session_dir()
    if not session_dir:
        return jsonify({"status": "error", "message": "Нет активной сессии"}), 400
    if not _has_consolidated_session_file(session_dir):
        return jsonify({"status": "error", "message": "Сводный прайс не найден"}), 400

    payload = _preexport_quality_payload_for_session(session_dir)
    if payload is None:
        return jsonify({"status": "error", "message": "Не удалось подготовить прайс для проверки"}), 500
    return jsonify(payload)


app.register_blueprint(create_export_bp(
    download=download,
    export_google_sheets=api_export_google_sheets,
    download_id_quality_report=download_id_quality_report,
    download_id_compare_report=download_id_compare_report,
    preexport_quality_check=api_preexport_quality_check,
))


@_serialized_price_mutation
def api_reapply_saved_markups():
    result = _get_category_extra_runtime().reapply_saved_markups()
    if isinstance(result, tuple):
        payload, status = result
        return jsonify(payload), status
    return jsonify(result)


app.register_blueprint(create_category_management_bp(
    name="category_management_extra_api",
    handlers={
        "/api/category-markups": (api_category_markups, ("GET",)),
        "/api/category-override-bulk": (api_category_override_bulk, ("POST",)),
        "/api/category-autosort-preview": (api_category_autosort_preview, ("POST",)),
        "/api/category-autosort-apply": (api_category_autosort_apply, ("POST",)),
        "/api/reapply-saved-markups": (api_reapply_saved_markups, ("POST",)),
    }
))


if __name__ == "__main__":
    try:
        local_port = int(os.getenv("PRICE_MIXER_PORT", "5001"))
    except ValueError:
        raise SystemExit("PRICE_MIXER_PORT must be an integer")
    if not 1 <= local_port <= 65535:
        raise SystemExit("PRICE_MIXER_PORT must be between 1 and 65535")
    print("=" * 50)
    print("Price Mixer Web")
    print(f"Открой в браузере: http://localhost:{local_port}")
    print("=" * 50)
    init_onliner_db()
    # The application emits its own access log without query strings.
    logging.getLogger("werkzeug").setLevel(logging.ERROR)
    # Stable local run mode: no Flask reloader double-process.
    app.run(
        debug=False,
        use_reloader=False,
        load_dotenv=False,
        host="127.0.0.1",
        port=local_port,
    )
